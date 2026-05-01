"""
╔══════════════════════════════════════════════════════════════════════════════╗
║  ORACLE V36.0 — MERGED ENTERPRISE EDITION                                    ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  Περιεχόμενα:                                                               ║
║  • Πλήρες μοντέλο V35.2 (xG blending, Pythagorean, Form PPG, GK adj,       ║
║    team totals factor, dynamic home advantage, σωστά totals/gk στο mu)     ║
║  • ΝΕΑ V36:                                                                 ║
║    - Auto-Calibration (Brier score από ledger)                             ║
║    - Fatigue Penalty (ξεκούραση < 96h → -15%)                             ║
║    - Hedging Engine (Dutching calculator)                                  ║
║    - Ledger-based confidence adjustment                                    ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

import csv
import hashlib
import io
import json
import logging
import math
import os
import sqlite3
import time
from datetime import datetime, timedelta, timezone
from typing import Optional, Dict, List, Tuple

import requests
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

try:
    from apscheduler.schedulers.blocking import BlockingScheduler
    SCHED_OK = True
except ImportError:
    SCHED_OK = False

# ══════════════════════════════════════════════════════════════════════════════
# 2. CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════
ODDS_API_KEY     = os.getenv("ODDS_API_KEY", "44e284d6037317389a0d873bd633d546")
API_FOOTBALL_KEY = os.getenv("API_FOOTBALL_KEY", "57500312c70c8c45a39faecd6942f3ea")
TELEGRAM_TOKEN   = os.getenv("TELEGRAM_TOKEN", "")
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID", "")

if not ODDS_API_KEY or not API_FOOTBALL_KEY:
    raise RuntimeError("Missing required environment variables: ODDS_API_KEY and API_FOOTBALL_KEY")

EV_THRESHOLD      = 1.04
PATTERN_THRESHOLD = 60
TRAP_MARGIN       = 18
KELLY_FRACTION    = 0.25
MAX_KELLY_PCT     = 0.03
BANKROLL          = float(os.getenv("BANKROLL", "1000"))

UTC_OFFSET_HOURS = 3
LOOKAHEAD_DAYS   = 3
LAST_N_FIXTURES  = 10
REQUEST_DELAY    = 0.5
XG_SAMPLE_SIZE   = 8
PRE_FILTER_EV    = 1.01

ELO_K            = 32
DEFAULT_ELO      = 1500.0
ELO_HOME_BONUS   = 50.0
RHO_CORRELATION  = -0.10
CONFIDENCE_HIGH  = 0.70
CONFIDENCE_MED   = 0.45
CACHE_TEAM_HOURS = 96
CACHE_XG_HOURS   = 96
CACHE_REF_DAYS   = 14
CACHE_SCORER_DAYS= 7

SHARP_BOOKMAKERS = ["pinnacle"]
PENALTY_XG_VALUE = 0.79

DB_FILE     = "oracle.db"
OUTPUT_FILE = "Oracle_V36_Enterprise.xlsx"

LEAGUE_BASELINE_DAYS = 7
HOME_CORNER_ADV = 1.12

# --- V36: Fatigue & Calibration Parameters ---
FATIGUE_LIMIT_H    = 96       # 4 ημέρες ξεκούρασης [16]
BRIER_SAMPLES      = 20       # Δείγμα Ledger για auto-calibration [17]
LEDGER_FILE        = "Oracle_Historical_Ledger.xlsx"

# --- Ensemble weights (V35.2) ---
W_DC_XG        = 0.40
W_ELO          = 0.20
W_MARKET       = 0.20
W_PYTHAGOREAN  = 0.10
W_FORM         = 0.10

# --- League‑specific Pythagorean exponents ---
LEAGUE_PYTH_EXP: Dict[str, float] = {
    "soccer_epl": 1.35,
    "soccer_efl_champ": 1.45,
    "soccer_england_league1": 1.55,
    "soccer_spain_la_liga": 1.40,
    "soccer_spain_segunda_division": 1.50,
    "soccer_italy_serie_a": 1.55,
    "soccer_italy_serie_b": 1.65,
    "soccer_germany_bundesliga": 1.45,
    "soccer_germany_2_bundesliga": 1.55,
    "soccer_france_ligue_one": 1.50,
    "soccer_france_ligue_two": 1.60,
    "soccer_portugal_primeira_liga": 1.50,
    "soccer_netherlands_eredivisie": 1.40,
    "soccer_greece_super_league": 1.60,
    "soccer_greece_super_league_2": 1.70,
    "soccer_turkey_super_league": 1.55,
    "soccer_sweden_allsvenskan": 1.65,
    "soccer_norway_eliteserien": 1.60,
    "soccer_denmark_superliga": 1.60,
    "soccer_uefa_champs_league": 1.40,
    "soccer_uefa_europa_league": 1.45,
    "soccer_uefa_europa_conference_league": 1.50,
}

# --- xG blending thresholds ---
XG_FULL_BLEND_SAMPLE  = 12
XG_HIGH_BLEND_SAMPLE  = 6
XG_MED_BLEND_SAMPLE   = 3

# --- Form momentum weights ---
FORM_WEIGHTS = [0.35, 0.25, 0.20, 0.12, 0.08]

# --- Dynamic home advantage ---
DYN_HOME_GD_BOOST_PER_GOAL = 0.05
DYN_HOME_GD_PENALTY_PER_GOAL = 0.05

FEATURES = {
    "elo": True, "poisson": True, "xg": True, "injuries": True,
    "weather": True, "referee": True, "h2h": True, "arbitrage": True,
    "telegram": bool(TELEGRAM_TOKEN and TELEGRAM_CHAT_ID),
    "scheduler": False, "backtest": False,
    "pythagorean": True,
    "form_momentum": True,
    "dynamic_home": True,
    "xg_poisson": True,
    "team_totals_factor": True,
    "goalkeeper_adjustment": True,
    "fatigue": True,         # V36
    "auto_calibration": True, # V36
    "hedging": True,          # V36
}

HOME_ADV: Dict[str, float] = {
    "soccer_epl": 1.32, "soccer_efl_champ": 1.28, "soccer_england_league1": 1.25,
    "soccer_spain_la_liga": 1.30, "soccer_spain_segunda_division": 1.24,
    "soccer_italy_serie_a": 1.28, "soccer_italy_serie_b": 1.22,
    "soccer_germany_bundesliga": 1.25, "soccer_germany_2_bundesliga": 1.22,
    "soccer_france_ligue_one": 1.27, "soccer_france_ligue_two": 1.22,
    "soccer_portugal_primeira_liga": 1.30, "soccer_netherlands_eredivisie": 1.28,
    "soccer_greece_super_league": 1.35, "soccer_greece_super_league_2": 1.33,
    "soccer_turkey_super_league": 1.35, "soccer_sweden_allsvenskan": 1.20,
    "soccer_norway_eliteserien": 1.18, "soccer_denmark_superliga": 1.22,
    "soccer_uefa_champs_league": 1.15, "soccer_uefa_europa_league": 1.12,
    "soccer_uefa_europa_conference_league": 1.10,
}

LEAGUE_STRENGTH: Dict[str, float] = {
    "Premier League": 1.00, "Serie A": 0.95, "Bundesliga": 0.95,
    "La Liga": 0.93, "Ligue 1": 0.90, "Championship": 0.85,
    "League One": 0.75, "Segunda División": 0.75, "Serie B": 0.75,
    "Ligue 2": 0.75, "Primeira Liga": 0.75, "Eredivisie": 0.72,
    "Super Lig": 0.70, "Super League 1": 0.65, "Superliga": 0.65,
    "Allsvenskan": 0.60, "Eliteserien": 0.60,
}

UEFA_WEIGHT: Dict[str, float] = {
    "soccer_uefa_champs_league": 1.20, "soccer_uefa_europa_league": 1.10,
    "soccer_uefa_europa_conference_league": 1.05,
}

LEAGUE_KEYS = list(HOME_ADV.keys())
MARKET_LABELS = ["1", "X", "2", "Over_2.5", "Under_2.5", "BTTS", "DNB_1", "DNB_2", "1X", "X2", "1_Over", "2_Over"]

logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(levelname)-8s  %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger(__name__)

def _parse_ts(ts: str) -> datetime:
    dt = datetime.fromisoformat(ts)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt

# ══════════════════════════════════════════════════════════════════════════════
# 3. DATABASE
# ══════════════════════════════════════════════════════════════════════════════
def get_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    _init_schema(conn)
    return conn

def _init_schema(conn: sqlite3.Connection) -> None:
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS elo_ratings (team_id INTEGER PRIMARY KEY, team_name TEXT NOT NULL, rating REAL NOT NULL DEFAULT 1500, updated TEXT);
    CREATE TABLE IF NOT EXISTS team_cache (team_name TEXT PRIMARY KEY, team_id INTEGER, league TEXT, fixtures TEXT, xg_for REAL, xg_against REAL, updated TEXT);
    CREATE TABLE IF NOT EXISTS referee_cache (name TEXT PRIMARY KEY, avg_cards REAL, avg_pens REAL, updated TEXT);
    CREATE TABLE IF NOT EXISTS picks (id TEXT PRIMARY KEY, run_ts TEXT, match_date TEXT, match_time TEXT, league TEXT, match TEXT, market TEXT, odds REAL, stat_p REAL, ev REAL, kelly REAL, pick_type TEXT, trend TEXT, result TEXT, pl REAL, close_odds REAL, clv_pct REAL);
    CREATE TABLE IF NOT EXISTS arb_log (id TEXT PRIMARY KEY, logged_ts TEXT, match TEXT, margin REAL, legs TEXT);
    CREATE TABLE IF NOT EXISTS odds_history (key TEXT PRIMARY KEY, price REAL NOT NULL, updated TEXT NOT NULL);
    CREATE TABLE IF NOT EXISTS team_stats_v33 (team_id INTEGER PRIMARY KEY, corners_for REAL, corners_agt REAL, cards_for REAL, cards_agt REAL, updated TEXT);
    CREATE TABLE IF NOT EXISTS scorer_cache (league_id INTEGER, season INTEGER, team_id INTEGER, player_name TEXT, goals INTEGER, matches INTEGER, updated TEXT, PRIMARY KEY(league_id, team_id));
    CREATE TABLE IF NOT EXISTS league_baseline (league_key TEXT PRIMARY KEY, avg_goals REAL NOT NULL, sample_n INTEGER, updated TEXT);
    CREATE TABLE IF NOT EXISTS team_totals (team_id INTEGER PRIMARY KEY, over25_freq REAL, avg_total_goals REAL, updated TEXT);
    CREATE TABLE IF NOT EXISTS goalkeeper_cache (team_id INTEGER PRIMARY KEY, save_pct REAL, updated TEXT);
    CREATE INDEX IF NOT EXISTS idx_team_name ON team_cache(team_name);
    CREATE INDEX IF NOT EXISTS idx_odds_key ON odds_history(key);
    """)
    conn.commit()

# ══════════════════════════════════════════════════════════════════════════════
# 3b. PER-LEAGUE GOAL BASELINE
# ══════════════════════════════════════════════════════════════════════════════
def get_league_baseline(conn: sqlite3.Connection, league_key: str) -> float:
    row = conn.execute(
        "SELECT avg_goals, updated FROM league_baseline WHERE league_key=?", (league_key,)
    ).fetchone()
    if row and row["updated"]:
        age_d = (datetime.now(timezone.utc) - _parse_ts(row["updated"])).days
        if age_d < LEAGUE_BASELINE_DAYS:
            return row["avg_goals"]
    return 2.60

def update_league_baseline(conn: sqlite3.Connection, league_key: str, fixtures: list) -> None:
    totals = [
        m["goals"]["home"] + m["goals"]["away"]
        for m in fixtures
        if m["goals"]["home"] is not None and m["goals"]["away"] is not None
    ]
    if len(totals) < 3:
        return
    avg = round(sum(totals) / len(totals), 4)
    conn.execute(
        "INSERT OR REPLACE INTO league_baseline VALUES (?,?,?,?)",
        (league_key, avg, len(totals), datetime.now(timezone.utc).isoformat())
    )
    conn.commit()

# ══════════════════════════════════════════════════════════════════════════════
# 4. ELO RATING SYSTEM
# ══════════════════════════════════════════════════════════════════════════════
def elo_expected_score(r_a: float, r_b: float) -> float:
    return 1.0 / (1.0 + 10.0 ** ((r_b - r_a) / 400.0))

def get_elo(conn: sqlite3.Connection, team_id: int) -> float:
    row = conn.execute("SELECT rating FROM elo_ratings WHERE team_id=?", (team_id,)).fetchone()
    return row["rating"] if row else DEFAULT_ELO

def update_elo_from_result(conn: sqlite3.Connection, home_id: int, home_name: str, home_goals: int,
                           away_id: int, away_name: str, away_goals: int) -> None:
    r_h, r_a = get_elo(conn, home_id), get_elo(conn, away_id)
    exp_h = elo_expected_score(r_h + ELO_HOME_BONUS, r_a)
    actual_h = 1.0 if home_goals > away_goals else (0.5 if home_goals == away_goals else 0.0)
    delta = ELO_K * (actual_h - exp_h)
    now = datetime.now(timezone.utc).isoformat()
    conn.execute("INSERT OR REPLACE INTO elo_ratings VALUES (?,?,?,?)", (home_id, home_name, round(r_h + delta, 2), now))
    conn.execute("INSERT OR REPLACE INTO elo_ratings VALUES (?,?,?,?)", (away_id, away_name, round(r_a - delta, 2), now))
    conn.commit()

def elo_probabilities(elo_h: float, elo_a: float) -> Tuple[float, float, float]:
    exp_h = elo_expected_score(elo_h + ELO_HOME_BONUS, elo_a)
    gap = abs(elo_h - elo_a)
    p_draw = max(0.12, 0.30 - gap * 0.00035)
    p1 = max(0.05, exp_h - p_draw / 2.0)
    p2 = max(0.05, 1 - exp_h - p_draw / 2.0)
    px = max(0.05, 1.0 - p1 - p2)
    t = p1 + px + p2
    return p1/t, px/t, p2/t

# ══════════════════════════════════════════════════════════════════════════════
# 5. DATA FETCHERS
# ══════════════════════════════════════════════════════════════════════════════
def _apf(endpoint: str, params: dict = {}) -> Optional[dict]:
    time.sleep(REQUEST_DELAY)
    try:
        r = requests.get(f"https://v3.football.api-sports.io{endpoint}",
                         headers={"x-apisports-key": API_FOOTBALL_KEY},
                         params=params, timeout=12)
        return r.json()
    except requests.RequestException as e:
        log.debug(f"API-Football {endpoint}: {e}")
        return None

def fetch_team_data(conn: sqlite3.Connection, team_name: str, league_key: str = "") -> Tuple[Optional[list], Optional[int], str, Optional[int], Optional[int]]:
    row = conn.execute("SELECT team_id, league, fixtures, updated FROM team_cache WHERE team_name=?", (team_name,)).fetchone()
    if row and row["updated"]:
        if (datetime.now(timezone.utc) - _parse_ts(row["updated"])).total_seconds() / 3600 < CACHE_TEAM_HOURS:
            fix = json.loads(row["fixtures"])
            l_id = fix[0]["league"]["id"] if fix else None
            season = fix[0]["league"]["season"] if fix else None
            return fix, row["team_id"], row["league"], l_id, season

    for name in [team_name, team_name.split()[0]]:
        res = _apf("/teams", {"search": name})
        if not res or not res.get("response"):
            continue
        t_id = res["response"][0]["team"]["id"]
        fix = _apf("/fixtures", {"team": t_id, "last": LAST_N_FIXTURES})
        if not fix or not fix.get("response"):
            continue
        fixtures = fix["response"]
        league = fixtures[0]["league"]["name"]
        l_id = fixtures[0]["league"]["id"]
        season = fixtures[0]["league"]["season"]

        for m in fixtures:
            gh, ga = m["goals"]["home"], m["goals"]["away"]
            if gh is not None and ga is not None:
                update_elo_from_result(conn, m["teams"]["home"]["id"], m["teams"]["home"]["name"],
                                       gh, m["teams"]["away"]["id"], m["teams"]["away"]["name"], ga)

        if league_key:
            update_league_baseline(conn, league_key, fixtures)

        conn.execute("INSERT OR REPLACE INTO team_cache (team_name,team_id,league,fixtures,updated) VALUES (?,?,?,?,?)",
                     (team_name, t_id, league, json.dumps(fixtures), datetime.now(timezone.utc).isoformat()))
        conn.commit()
        return fixtures, t_id, league, l_id, season
    return None, None, "Unknown", None, None

def fetch_advanced_stats(conn: sqlite3.Connection, team_name: str, team_id: int) -> Optional[dict]:
    row = conn.execute("SELECT xg_for, xg_against, updated FROM team_cache WHERE team_name=?", (team_name,)).fetchone()
    stats_row = conn.execute("SELECT * FROM team_stats_v33 WHERE team_id=?", (team_id,)).fetchone()

    if row and stats_row and row["updated"]:
        if (datetime.now(timezone.utc) - _parse_ts(row["updated"])).total_seconds() / 3600 < CACHE_XG_HOURS:
            return {
                "xg_for": row["xg_for"], "xg_against": row["xg_against"],
                "corn_for": stats_row["corners_for"], "corn_agt": stats_row["corners_agt"],
                "card_for": stats_row["cards_for"], "card_agt": stats_row["cards_agt"],
                "xg_sample_count": XG_SAMPLE_SIZE
            }

    fix_res = _apf("/fixtures", {"team": team_id, "last": XG_SAMPLE_SIZE})
    if not fix_res or not fix_res.get("response"):
        return None

    xgf = xga = corn_f = corn_a = card_f = card_a = 0.0
    count = 0
    xg_sample_count = 0

    for match in fix_res["response"]:
        fix_id = match["fixture"]["id"]
        opp_id = match["teams"]["away"]["id"] if match["teams"]["home"]["id"] == team_id else match["teams"]["home"]["id"]
        diff_factor = max(0.8, min(1.2, get_elo(conn, opp_id) / 1500.0))

        stats = _apf("/fixtures/statistics", {"fixture": fix_id})
        if not stats or not stats.get("response"):
            continue

        raw_xgf = 0.0
        for ts in stats["response"]:
            if ts["team"]["id"] == team_id:
                for s in ts.get("statistics", []):
                    if s["type"] == "expected_goals" and s["value"] is not None:
                        try:
                            raw_xgf = float(s["value"])
                        except ValueError:
                            pass

        pens_for = pens_against = 0
        if raw_xgf > 2.5:
            events = _apf("/fixtures/events", {"fixture": fix_id})
            if events and events.get("response"):
                for e in events["response"]:
                    if e.get("type") == "Goal" and e.get("detail") == "Penalty":
                        if e["team"]["id"] == team_id:
                            pens_for += 1
                        else:
                            pens_against += 1

        fixture_counted = False
        has_xg_this_match = False
        for ts in stats["response"]:
            is_us = ts["team"]["id"] == team_id
            for s in ts.get("statistics", []):
                val = s["value"]
                if val is None:
                    continue
                if s["type"] not in ("expected_goals", "Corner Kicks", "Yellow Cards", "Red Cards"):
                    continue
                try:
                    val = float(val)
                except ValueError:
                    continue

                if s["type"] == "expected_goals":
                    has_xg_this_match = True
                    if is_us:
                        xgf += max(0.0, val - (pens_for * PENALTY_XG_VALUE)) * diff_factor
                    else:
                        xga += max(0.0, val - (pens_against * PENALTY_XG_VALUE)) / diff_factor
                elif s["type"] == "Corner Kicks":
                    if is_us:
                        corn_f += val
                    else:
                        corn_a += val
                elif s["type"] in ("Yellow Cards", "Red Cards"):
                    w = 1 if s["type"] == "Yellow Cards" else 2
                    if is_us:
                        card_f += val * w
                    else:
                        card_a += val * w

            if is_us and not fixture_counted:
                count += 1
                fixture_counted = True
                if has_xg_this_match:
                    xg_sample_count += 1

    if count == 0:
        return None
    res = {
        "xg_for": round(xgf/count, 3),
        "xg_against": round(xga/count, 3),
        "corn_for": round(corn_f/count, 2),
        "corn_agt": round(corn_a/count, 2),
        "card_for": round(card_f/count, 2),
        "card_agt": round(card_a/count, 2),
        "xg_sample_count": xg_sample_count
    }
    conn.execute("UPDATE team_cache SET xg_for=?, xg_against=?, updated=? WHERE team_name=?",
                 (res["xg_for"], res["xg_against"], datetime.now(timezone.utc).isoformat(), team_name))
    conn.execute("INSERT OR REPLACE INTO team_stats_v33 VALUES (?,?,?,?,?,?)",
                 (team_id, res["corn_for"], res["corn_agt"], res["card_for"], res["card_agt"],
                  datetime.now(timezone.utc).isoformat()))
    conn.commit()
    return res

def fetch_team_totals_factor(conn: sqlite3.Connection, team_id: int, fixtures: list) -> float:
    if not FEATURES["team_totals_factor"]:
        return 1.0

    row = conn.execute("SELECT over25_freq, updated FROM team_totals WHERE team_id=?", (team_id,)).fetchone()
    if row and row["updated"]:
        if (datetime.now(timezone.utc) - _parse_ts(row["updated"])).days < 7:
            return row["over25_freq"]

    over_count = 0
    total_matches = 0
    for m in fixtures:
        gh = m["goals"]["home"]
        ga = m["goals"]["away"]
        if gh is not None and ga is not None:
            total_matches += 1
            if gh + ga > 2.5:
                over_count += 1
    if total_matches < 3:
        return 1.0
    freq = over_count / total_matches
    league_avg = 0.50
    factor = freq / league_avg
    factor = max(0.75, min(1.25, factor))

    conn.execute("INSERT OR REPLACE INTO team_totals VALUES (?,?,?,?)",
                 (team_id, factor, None, datetime.now(timezone.utc).isoformat()))
    conn.commit()
    return factor

def fetch_goalkeeper_adjustment(conn: sqlite3.Connection, team_id: int, fixtures: list) -> float:
    if not FEATURES["goalkeeper_adjustment"]:
        return 1.0

    row = conn.execute("SELECT save_pct, updated FROM goalkeeper_cache WHERE team_id=?", (team_id,)).fetchone()
    if row and row["updated"]:
        if (datetime.now(timezone.utc) - _parse_ts(row["updated"])).days < 14:
            return row["save_pct"]

    LEAGUE_AVG_SAVE_PCT = 0.67
    goals_conceded = 0
    count = 0

    for m in fixtures[:LAST_N_FIXTURES]:
        gh = m["goals"]["home"]
        ga = m["goals"]["away"]
        if gh is None or ga is None:
            continue
        is_home = m["teams"]["home"]["id"] == team_id
        conceded = ga if is_home else gh
        goals_conceded += conceded
        count += 1

    if count < 3:
        return 1.0

    estimated_shots = goals_conceded / max(1 - LEAGUE_AVG_SAVE_PCT, 0.10)
    if estimated_shots == 0:
        return 1.0

    team_save_pct = 1.0 - (goals_conceded / estimated_shots)
    adjustment = round(LEAGUE_AVG_SAVE_PCT / max(team_save_pct, 0.30), 3)
    adjustment = max(0.85, min(1.15, adjustment))

    conn.execute("INSERT OR REPLACE INTO goalkeeper_cache VALUES (?,?,?)",
                 (team_id, adjustment, datetime.now(timezone.utc).isoformat()))
    conn.commit()
    return adjustment

def fetch_top_scorer(conn: sqlite3.Connection, league_id: int, season: int, team_id: int) -> Optional[dict]:
    if not league_id or not season:
        return None
    row = conn.execute("SELECT player_name, goals, matches, updated FROM scorer_cache WHERE league_id=? AND team_id=?",
                       (league_id, team_id)).fetchone()
    if row and row["updated"]:
        if (datetime.now(timezone.utc) - _parse_ts(row["updated"])).days < CACHE_SCORER_DAYS:
            return {"name": row["player_name"], "goals": row["goals"], "matches": row["matches"]}

    log.info(f"      📥 Fetching Top Scorers for League {league_id}...")
    data = _apf("/players/topscorers", {"league": league_id, "season": season})
    if not data or not data.get("response"):
        return None

    now = datetime.now(timezone.utc).isoformat()
    best_scorers = {}
    for p in data["response"]:
        p_name = p["player"]["name"]
        for stat in p["statistics"]:
            t_id = stat["team"]["id"]
            goals = stat["goals"]["total"] or 0
            matches = stat["games"]["appearences"] or 1
            if t_id not in best_scorers or goals > best_scorers[t_id]["goals"]:
                best_scorers[t_id] = {"name": p_name, "goals": goals, "matches": matches}

    for t_id, info in best_scorers.items():
        conn.execute("INSERT OR REPLACE INTO scorer_cache VALUES (?,?,?,?,?,?,?)",
                     (league_id, season, t_id, info["name"], info["goals"], info["matches"], now))
    conn.commit()
    return best_scorers.get(team_id)

def fetch_injuries(team_id: int, fixtures: list) -> float:
    if not FEATURES["injuries"] or not fixtures:
        return 0.0
    last_fid = fixtures[-1]["fixture"]["id"]
    data = _apf("/fixtures/players", {"fixture": last_fid})
    if not data or not data.get("response"):
        return 0.0

    impact = 0.0
    for team_stat in data["response"]:
        if team_stat["team"]["id"] == team_id:
            for p in team_stat.get("players", []):
                stats = p.get("statistics", [{}])[0]
                games = stats.get("games", {})
                if games.get("minutes") is None:
                    rating = float(games.get("rating") or 6.0)
                    if rating > 6.0:
                        impact += (rating - 6.0)
            return min(impact, 5.0)
    return 0.0

def fetch_referee_profile(conn: sqlite3.Connection, name: str) -> dict:
    if not FEATURES["referee"] or not name:
        return {"cards_mod": 1.0, "pens_mod": 1.0}
    row = conn.execute("SELECT avg_cards, avg_pens, updated FROM referee_cache WHERE name=?", (name,)).fetchone()
    if row and row["updated"]:
        if (datetime.now(timezone.utc) - _parse_ts(row["updated"])).days < 7:
            return {"cards_mod": row["avg_cards"] / 4.0, "pens_mod": row["avg_pens"] / 0.25}

    data = _apf("/fixtures", {"referee": name, "last": 20})
    if not data or not data.get("response"):
        return {"cards_mod": 1.0, "pens_mod": 1.0}

    total_cards = total_pens = count = 0
    for m in data["response"]:
        evts = _apf("/fixtures/events", {"fixture": m["fixture"]["id"]})
        if not evts or not evts.get("response"):
            continue
        for e in evts["response"]:
            if e.get("type") == "Card":
                total_cards += 1
            if e.get("detail") == "Penalty":
                total_pens += 1
        count += 1

    if count == 0:
        return {"cards_mod": 1.0, "pens_mod": 1.0}
    avg_c, avg_p = total_cards / count, total_pens / count
    conn.execute("INSERT OR REPLACE INTO referee_cache VALUES (?,?,?,?)",
                 (name, avg_c, avg_p, datetime.now(timezone.utc).isoformat()))
    conn.commit()
    return {"cards_mod": avg_c / 4.0, "pens_mod": avg_p / 0.25}

def fetch_h2h(conn: sqlite3.Connection, home_id: int, away_id: int) -> Optional[list]:
    if not FEATURES["h2h"] or not home_id or not away_id:
        return None
    data = _apf("/fixtures/headtohead", {"h2h": f"{home_id}-{away_id}", "last": 5})
    return data.get("response") if data else None

# ══════════════════════════════════════════════════════════════════════════════
# 6. V36 NEW FEATURES: CALIBRATION, FATIGUE, HEDGING
# ══════════════════════════════════════════════════════════════════════════════

# --- 6a. Auto-Calibration from Ledger ---
def get_brier_calibration(market: str) -> float:
    """
    Διαβάζει το ιστορικό αρχείο Oracle_Historical_Ledger.xlsx
    και υπολογίζει τον συντελεστή διόρθωσης βάσει Brier score.
    Επιστρέφει multiplier για το stat_p (π.χ. 1.15 = μείωση 13%).
    """
    if not FEATURES["auto_calibration"]:
        return 1.0
    try:
        if not os.path.exists(LEDGER_FILE):
            return 1.0
        df = pd.read_excel(LEDGER_FILE, engine='openpyxl')
        if 'Market' not in df.columns or 'Verdict_Code' not in df.columns:
            return 1.0
        recent = df[df['Market'] == market].tail(BRIER_SAMPLES)
        if len(recent) < 10:
            return 1.0
        hit_rate = (recent['Verdict_Code'] == 'CORRECT').mean()
        # Εμπειρικός κανόνας: χαμηλό hit rate → μείωση πιθανότητας
        if hit_rate < 0.40:
            return 1.30  # Μεγάλη διόρθωση
        elif hit_rate < 0.50:
            return 1.15  # Μέτρια διόρθωση
        return 1.0
    except Exception as e:
        log.debug(f"Brier calibration skipped: {e}")
        return 1.0

# --- 6b. Fatigue Penalty ---
def apply_fatigue_penalty(team_id: int, fixtures: list, match_time: datetime) -> float:
    """
    Αν η ομάδα έπαιξε < FATIGUE_LIMIT_H ώρες πριν, εφαρμόζει ποινή -15%.
    """
    if not FEATURES["fatigue"] or not fixtures:
        return 1.0
    last_match = None
    for f in reversed(fixtures):
        if f['fixture']['status']['short'] == 'FT':
            last_match = datetime.fromisoformat(f['fixture']['date'].replace('Z', '+00:00'))
            break
    if last_match:
        diff_h = (match_time - last_match).total_seconds() / 3600
        if diff_h < FATIGUE_LIMIT_H:
            return 0.85  # 15% Fatigue Penalty
    return 1.0

# --- 6c. Hedging Engine ---
def calculate_hedge_note(market: str, odds_1: float, odds_x: float, odds_2: float) -> str:
    """
    Dutching calculator για προστασία στοιχήματος.
    Επιστρέφει προτεινόμενο πλάνο αντιστάθμισης.
    """
    if market not in ('1', '2'):
        return ""
    unit = 5.0  # συνολικό ποντάρισμα
    if market == '1':
        stake_draw = round(unit / odds_x, 2) if odds_x else 0
        rem = unit - stake_draw
        stake_win = round(rem * 0.65, 2)
        stake_combo = round(unit - stake_draw - stake_win, 2)
        return f"X: €{stake_draw} | 1: €{stake_win} | 1&Ov: €{stake_combo}"
    else:  # market == '2'
        stake_draw = round(unit / odds_x, 2) if odds_x else 0
        rem = unit - stake_draw
        stake_win = round(rem * 0.65, 2)
        stake_combo = round(unit - stake_draw - stake_win, 2)
        return f"X: €{stake_draw} | 2: €{stake_win} | 2&Ov: €{stake_combo}"

# ══════════════════════════════════════════════════════════════════════════════
# 7. STATISTICAL MODEL (V35.2 + V36 adjustments)
# ══════════════════════════════════════════════════════════════════════════════

# --- Pythagorean ---
def pythagorean_win_prob(gf: float, ga: float, league_key: str) -> float:
    exp = LEAGUE_PYTH_EXP.get(league_key, 1.50)
    gf = max(gf, 0.10)
    ga = max(ga, 0.10)
    return gf**exp / (gf**exp + ga**exp)

def pythagorean_match_probs(h_gf: float, h_ga: float, a_gf: float, a_ga: float,
                            league_key: str, home_adv: float = 1.28) -> Tuple[float, float, float]:
    ph = pythagorean_win_prob(h_gf * home_adv, h_ga, league_key)
    pa = pythagorean_win_prob(a_gf, a_ga * home_adv, league_key)

    denom = ph + pa - 2 * ph * pa
    if denom <= 0:
        p1 = 0.45
        p2 = 0.30
    else:
        p1 = max(0.05, min(0.85, (ph - ph * pa) / denom))
        p2 = max(0.05, min(0.85, (pa - ph * pa) / denom))

    gap = abs(ph - pa)
    p_draw = max(0.10, 0.28 - gap * 0.30)
    px = max(0.05, 1.0 - p1 - p2)
    px = min(px, p_draw)

    t = p1 + px + p2
    return p1/t, px/t, p2/t

# --- Form Momentum ---
def form_momentum_ppg(fixtures: list, team_id: int) -> float:
    results = []
    for m in reversed(fixtures[:LAST_N_FIXTURES]):
        gh = m["goals"]["home"]
        ga = m["goals"]["away"]
        if gh is None or ga is None:
            continue
        is_home = m["teams"]["home"]["id"] == team_id
        scored, conceded = (gh, ga) if is_home else (ga, gh)
        if scored > conceded:
            results.append(3)
        elif scored == conceded:
            results.append(1)
        else:
            results.append(0)

    if not results:
        return 0.50

    total_weight = 0.0
    weighted_sum = 0.0
    for i, pts in enumerate(results[:5]):
        w = FORM_WEIGHTS[i] if i < len(FORM_WEIGHTS) else 0.05
        weighted_sum += pts * w
        total_weight += w

    if total_weight == 0:
        return 0.50
    ppg = weighted_sum / total_weight
    momentum_p = 0.50 + (ppg / 3.0) * 0.35
    return min(0.85, max(0.15, momentum_p))

def form_match_probs(h_momentum: float, a_momentum: float) -> Tuple[float, float, float]:
    total = h_momentum + a_momentum
    if total <= 0:
        return 0.40, 0.28, 0.32
    p1_raw = h_momentum / total
    p2_raw = a_momentum / total
    diff = abs(h_momentum - a_momentum)
    p_draw = max(0.12, 0.30 - diff * 0.50)
    p1 = max(0.05, min(0.85, p1_raw * (1 - p_draw)))
    p2 = max(0.05, min(0.85, p2_raw * (1 - p_draw)))
    px = max(0.05, 1.0 - p1 - p2)
    t = p1 + px + p2
    return p1/t, px/t, p2/t

# --- Dynamic Home Advantage ---
def dynamic_home_advantage_gd(fixtures: list, team_id: int, base_ha: float) -> float:
    if not FEATURES["dynamic_home"]:
        return base_ha

    home_games = 0
    total_gd = 0
    for m in fixtures[:LAST_N_FIXTURES]:
        if m["teams"]["home"]["id"] != team_id:
            continue
        gh = m["goals"]["home"]
        ga = m["goals"]["away"]
        if gh is None or ga is None:
            continue
        home_games += 1
        total_gd += (gh - ga)

    if home_games < 3:
        return base_ha

    avg_gd = total_gd / home_games
    if avg_gd > 0:
        boost = min(0.20, avg_gd * DYN_HOME_GD_BOOST_PER_GOAL)
        return round(base_ha * (1.0 + boost), 3)
    elif avg_gd < 0:
        penalty = min(0.15, abs(avg_gd) * DYN_HOME_GD_PENALTY_PER_GOAL)
        return round(base_ha * (1.0 - penalty), 3)
    return base_ha

# --- xG‑Driven Poisson ---
def compute_lambda_mu(
    h_gf: float, h_ga: float,
    a_gf: float, a_ga: float,
    h_xg_for: Optional[float], h_xg_agt: Optional[float],
    a_xg_for: Optional[float], a_xg_agt: Optional[float],
    h_xg_sample: int, a_xg_sample: int,
    league_avg: float,
    home_f: float,
    h_totals_factor: float = 1.0,
    a_totals_factor: float = 1.0,
    h_gk_adj: float = 1.0,
    a_gk_adj: float = 1.0,
) -> Tuple[float, float, float, float]:
    la = max(league_avg, 0.50)

    lam_goals = max(0.20, (h_gf / la) * (a_ga / la) * la * home_f * h_totals_factor * h_gk_adj)
    mu_goals  = max(0.20, (a_gf / la) * (h_ga / la) * la / home_f * a_totals_factor * a_gk_adj)

    has_xg = (h_xg_for is not None and h_xg_agt is not None and
              a_xg_for is not None and a_xg_agt is not None and
              h_xg_for > 0 and a_xg_for > 0)

    if not has_xg or not FEATURES["xg_poisson"]:
        return lam_goals, mu_goals, lam_goals, mu_goals

    lam_xg = max(0.20, (h_xg_for / la) * (a_xg_agt / la) * la * home_f * h_totals_factor * h_gk_adj)
    mu_xg  = max(0.20, (a_xg_for / la) * (h_xg_agt / la) * la / home_f * a_totals_factor * a_gk_adj)

    min_sample = min(h_xg_sample, a_xg_sample)
    if min_sample >= XG_FULL_BLEND_SAMPLE:
        blend = 1.00
    elif min_sample >= XG_HIGH_BLEND_SAMPLE:
        blend = 0.85
    elif min_sample >= XG_MED_BLEND_SAMPLE:
        blend = 0.70
    else:
        blend = 0.50

    lam = round(lam_xg * blend + lam_goals * (1 - blend), 4)
    mu  = round(mu_xg  * blend + mu_goals  * (1 - blend), 4)

    return lam, mu, lam_xg, mu_xg

# --- Extended Poisson ---
def poisson_pmf(lam: float, k: int) -> float:
    if lam <= 0:
        return 1.0 if k == 0 else 0.0
    return math.exp(-lam) * (lam ** k) / math.factorial(k)

def poisson_over_prob(lam: float, line: float) -> float:
    k = int(math.floor(line))
    prob_under = sum(poisson_pmf(lam, i) for i in range(k + 1))
    return max(0.0, 1.0 - prob_under)

def bivariate_poisson_probs(lam: float, mu: float, rho: float = RHO_CORRELATION) -> Tuple[float, float, float, float, float]:
    p1 = px = p2 = p_ov = p_btts = 0.0
    max_goals = 15
    for i in range(max_goals + 1):
        for j in range(max_goals + 1):
            p = poisson_pmf(lam, i) * poisson_pmf(mu, j)
            if i == 0 and j == 0:
                adj = max(0.001, 1 - lam * mu * rho)
            elif i == 1 and j == 0:
                adj = 1 + mu * rho
            elif i == 0 and j == 1:
                adj = 1 + lam * rho
            elif i == 1 and j == 1:
                adj = 1 - rho
            else:
                adj = 1.0
            p = max(0.0, p * adj)
            if i > j:
                p1 += p
            elif i == j:
                px += p
            else:
                p2 += p
            if (i + j) > 2:
                p_ov += p
            if i >= 1 and j >= 1:
                p_btts += p

    total = p1 + px + p2
    if total == 0:
        total = 1.0
    p1 /= total
    px /= total
    p2 /= total
    p_ov = min(0.90, p_ov)
    p_btts = min(0.80, p_btts)
    return p1, px, p2, p_ov, p_btts

# --- Ensemble ---
def get_dynamic_weights(matches_played: float) -> Tuple[float, float]:
    if matches_played < 3:
        return 0.10, 0.70
    elif matches_played < 8:
        return 0.30, 0.50
    elif matches_played < 15:
        return 0.50, 0.30
    else:
        return 0.65, 0.15

def get_ensemble_stat_p(
    dc_p: float, elo_p: float, market_p: float,
    pyth_p: float, form_p: float,
    w_dc: float, w_elo: float
) -> Tuple[float, float, str]:
    dyn_intended = W_DC_XG + W_ELO
    if w_dc + w_elo > 0:
        scale = dyn_intended / (w_dc + w_elo)
    else:
        scale = 1.0

    eff_dc = w_dc * scale
    eff_elo = w_elo * scale
    eff_market = W_MARKET
    eff_pyth = W_PYTHAGOREAN
    eff_form = W_FORM

    total = eff_dc + eff_elo + eff_market + eff_pyth + eff_form
    eff_dc /= total
    eff_elo /= total
    eff_market /= total
    eff_pyth /= total
    eff_form /= total

    ensemble_p = (
        dc_p     * eff_dc +
        elo_p    * eff_elo +
        market_p * eff_market +
        pyth_p   * eff_pyth +
        form_p   * eff_form
    )
    ensemble_p = round(max(0.05, min(0.95, ensemble_p)), 4)

    models = [dc_p, elo_p, market_p, pyth_p, form_p]
    variance = sum((m - ensemble_p)**2 for m in models) / len(models)
    confidence = round(1.0 - min(variance * 8, 0.70), 4)

    if confidence >= CONFIDENCE_HIGH:
        agreement = "HIGH"
    elif confidence >= CONFIDENCE_MED:
        agreement = "MEDIUM"
    else:
        agreement = "LOW"

    return ensemble_p, confidence, agreement

def ema_form(fixtures: list, team_id: int, league_mod: float, alpha: float = 0.35) -> Tuple[float, float, float]:
    ema_val = None
    gf = ga = count = 0.0
    for m in reversed(fixtures[:LAST_N_FIXTURES]):
        gh, ga_g = m["goals"]["home"], m["goals"]["away"]
        if gh is None or ga_g is None:
            continue
        is_home = m["teams"]["home"]["id"] == team_id
        scored, conceded = (gh, ga_g) if is_home else (ga_g, gh)
        gf += scored
        ga += conceded
        count += 1
        result = (1.0 if scored > conceded else 0.5 if scored == conceded else 0.0) * league_mod
        ema_val = result if ema_val is None else alpha * result + (1 - alpha) * ema_val
    return (ema_val or 0.0), gf / max(count, 1), ga / max(count, 1)

def get_full_analysis(conn: sqlite3.Connection, home_name: str, away_name: str,
                      league_key: str, ref_name: str = "", match_utc: datetime = None) -> Optional[dict]:
    h_fix, h_id, h_league, l_id, season = fetch_team_data(conn, home_name, league_key)
    a_fix, a_id, a_league, _, _ = fetch_team_data(conn, away_name, league_key)
    if not h_fix or not a_fix:
        return None

    h_mod = LEAGUE_STRENGTH.get(h_league, 0.70)
    a_mod = LEAGUE_STRENGTH.get(a_league, 0.70)
    base_ha = HOME_ADV.get(league_key, 1.28)

    _, h_gf, h_ga = ema_form(h_fix, h_id, h_mod)
    _, a_gf, a_ga = ema_form(a_fix, a_id, a_mod)

    league_avg = get_league_baseline(conn, league_key)

    # Dynamic home advantage
    ha = dynamic_home_advantage_gd(h_fix, h_id, base_ha)
    home_f = ha ** 0.30

    # Form momentum
    h_momentum = form_momentum_ppg(h_fix, h_id)
    a_momentum = form_momentum_ppg(a_fix, a_id)
    form_p1, form_px, form_p2 = form_match_probs(h_momentum, a_momentum)

    # Advanced stats
    h_stats = fetch_advanced_stats(conn, home_name, h_id) if h_id else None
    a_stats = fetch_advanced_stats(conn, away_name, a_id) if a_id else None

    h_xg_for = h_stats["xg_for"] if h_stats else None
    h_xg_agt = h_stats["xg_against"] if h_stats else None
    a_xg_for = a_stats["xg_for"] if a_stats else None
    a_xg_agt = a_stats["xg_against"] if a_stats else None
    h_xg_sample = h_stats["xg_sample_count"] if h_stats else 0
    a_xg_sample = a_stats["xg_sample_count"] if a_stats else 0

    # Team totals & GK
    h_totals_factor = fetch_team_totals_factor(conn, h_id, h_fix) if h_id else 1.0
    a_totals_factor = fetch_team_totals_factor(conn, a_id, a_fix) if a_id else 1.0
    h_gk_adj = fetch_goalkeeper_adjustment(conn, h_id, h_fix) if h_id else 1.0
    a_gk_adj = fetch_goalkeeper_adjustment(conn, a_id, a_fix) if a_id else 1.0

    # V36: Fatigue penalty
    h_fatigue = apply_fatigue_penalty(h_id, h_fix, match_utc) if match_utc and h_id else 1.0
    a_fatigue = apply_fatigue_penalty(a_id, a_fix, match_utc) if match_utc and a_id else 1.0

    # Apply fatigue to λ/μ
    lam, mu, lam_xg, mu_xg = compute_lambda_mu(
        h_gf, h_ga, a_gf, a_ga,
        h_xg_for, h_xg_agt, a_xg_for, a_xg_agt,
        h_xg_sample, a_xg_sample,
        league_avg, home_f,
        h_totals_factor * h_fatigue, a_totals_factor * a_fatigue,
        h_gk_adj, a_gk_adj,
    )
    p1, px, p2, p_ov, p_btts = bivariate_poisson_probs(lam, mu)

    # Elo
    ep1 = epx = ep2 = None
    if FEATURES["elo"] and h_id and a_id:
        ep1, epx, ep2 = elo_probabilities(get_elo(conn, h_id), get_elo(conn, a_id))

    # Pythagorean
    pyth_p1 = pyth_px = pyth_p2 = None
    if FEATURES["pythagorean"] and h_gf > 0 and h_ga > 0 and a_gf > 0 and a_ga > 0:
        pyth_p1, pyth_px, pyth_p2 = pythagorean_match_probs(
            h_gf, h_ga, a_gf, a_ga, league_key, ha
        )

    h2h = fetch_h2h(conn, h_id, a_id)
    home_missing = fetch_injuries(h_id, h_fix) if FEATURES["injuries"] and h_id else 0.0
    away_missing = fetch_injuries(a_id, a_fix) if FEATURES["injuries"] and a_id else 0.0

    ref_profile = fetch_referee_profile(conn, ref_name) if ref_name else {"cards_mod": 1.0, "pens_mod": 1.0}

    # xG modifier
    if h_xg_for is not None and a_xg_for is not None:
        xg_ratio = h_xg_for / max(a_xg_for, 0.30)
        p1   = p1 * (1 + 0.07 * (xg_ratio - 1))
        p2   = p2 * (1 - 0.07 * (xg_ratio - 1))
        p_ov = max(0.10, min(0.90, p_ov * ((h_xg_for + a_xg_for) / 2.60)))

    # H2H modifier
    rivalry_mod = 1.0
    h2h_count = 0
    if h2h:
        wins = draws = losses = 0
        for m in h2h:
            gh, ga = m["goals"]["home"], m["goals"]["away"]
            if gh is None:
                continue
            sc, co = (gh, ga) if m["teams"]["home"]["id"] == h_id else (ga, gh)
            if sc > co:
                wins += 1
            elif sc == co:
                draws += 1
            else:
                losses += 1
            h2h_count += 1
        bias = (wins - losses) / max(wins + draws + losses, 1)
        p1 *= (1 + bias * 0.05)
        p2 *= (1 - bias * 0.05)
        rivalry_mod = 1.20 if h2h_count >= 3 else 1.0

    # Injury modifier
    p1 *= max(0.78, 1.0 - home_missing * 0.05)
    p2 *= max(0.78, 1.0 - away_missing * 0.05)

    # Renormalise
    p1 = max(0.05, min(0.85, p1))
    px = max(0.05, min(0.60, px))
    p2 = max(0.05, min(0.85, p2))
    t = p1 + px + p2
    p1, px, p2 = p1/t, px/t, p2/t

    # Specials
    specials = {"corners": "", "cards": "", "scorer": ""}

    if h_stats and a_stats:
        home_corners_exp = h_stats["corn_for"] * HOME_CORNER_ADV + a_stats["corn_agt"]
        away_corners_exp = a_stats["corn_for"] / HOME_CORNER_ADV + h_stats["corn_agt"]
        exp_corners = (home_corners_exp + away_corners_exp) / 2.0
        c_ov85 = poisson_over_prob(exp_corners, 8.5)
        c_ov95 = poisson_over_prob(exp_corners, 9.5)
        specials["corners"] = f"Exp: {exp_corners:.1f} | Ov8.5: {int(c_ov85*100)}% | Ov9.5: {int(c_ov95*100)}%"

        base_cards = (h_stats["card_for"] + a_stats["card_for"]) / 2.0
        exp_cards = base_cards * ref_profile["cards_mod"] * rivalry_mod
        card_ov35 = poisson_over_prob(exp_cards, 3.5)
        card_ov45 = poisson_over_prob(exp_cards, 4.5)
        rivalry_tag = " ⚔️Derby" if rivalry_mod > 1.0 else ""
        specials["cards"] = f"Exp: {exp_cards:.1f} | Ov3.5: {int(card_ov35*100)}% | Ov4.5: {int(card_ov45*100)}%{rivalry_tag}"

    h_scorer = fetch_top_scorer(conn, l_id, season, h_id)
    a_scorer = fetch_top_scorer(conn, l_id, season, a_id)
    best_scorer_txt = "N/A"
    best_prob = 0.0
    safe_la = max(league_avg, 0.10)
    for scorer, team_lam in [(h_scorer, lam), (a_scorer, mu)]:
        if scorer and scorer["matches"] > 0:
            goals_per_game = scorer["goals"] / scorer["matches"]
            prob = 1.0 - math.exp(-(goals_per_game * (team_lam / safe_la)))
            prob = max(0.0, min(0.99, prob))
            if prob > best_prob:
                best_prob = prob
                best_scorer_txt = f"{scorer['name']} ({int(prob*100)}%)"
    specials["scorer"] = best_scorer_txt

    avg_matches = (len(h_fix) + len(a_fix)) / 2.0

    xg_available = h_xg_for is not None and a_xg_for is not None
    log.debug(
        f"   📐 λ={lam:.3f}(xG={lam_xg:.3f}) μ={mu:.3f}(xG={mu_xg:.3f}) "
        f"ha={ha:.3f}(base={base_ha:.3f}) "
        f"h_ppg={h_momentum:.3f} a_ppg={a_momentum:.3f} "
        f"fatigue_h={h_fatigue:.2f} fatigue_a={a_fatigue:.2f} "
        f"pyth={'✓' if pyth_p1 else '✗'} xg={'✓' if xg_available else '✗'}"
    )

    return {
        "1": round(p1, 4), "X": round(px, 4), "2": round(p2, 4),
        "_dc_p1": p1, "_dc_px": px, "_dc_p2": p2,
        "_elo_p1": ep1, "_elo_px": epx, "_elo_p2": ep2,
        "_pyth_p1": pyth_p1, "_pyth_px": pyth_px, "_pyth_p2": pyth_p2,
        "_form_p1": form_p1, "_form_px": form_px, "_form_p2": form_p2,
        "_matches_played": avg_matches,
        "_h_momentum": h_momentum, "_a_momentum": a_momentum,
        "_lam": lam, "_mu": mu,
        "_ha_dynamic": ha,
        "Over_2.5": round(p_ov, 4), "Under_2.5": round(max(0.10, 1.0 - p_ov), 4),
        "BTTS": round(p_btts, 4),
        "DNB_1": round(p1 / (p1 + p2 or 1.0), 4),
        "DNB_2": round(p2 / (p1 + p2 or 1.0), 4),
        "1X": round(p1 + px, 4), "X2": round(px + p2, 4),
        "1_Over": round(p1 * p_ov, 4), "2_Over": round(p2 * p_ov, 4),
        "Specials_Corners": specials["corners"],
        "Specials_Cards": specials["cards"],
        "Specials_Scorer": specials["scorer"],
    }

# ══════════════════════════════════════════════════════════════════════════════
# 8. ODDS PARSING & VIG REMOVAL
# ══════════════════════════════════════════════════════════════════════════════
def parse_sharp_and_soft_odds(match: dict, home: str, away: str) -> Tuple[dict, dict, dict]:
    sharp, soft, per_book = {}, {}, {}
    for bk in match.get("bookmakers", []):
        bk_key = bk.get("key", "?")
        per_book.setdefault(bk_key, {})
        is_sharp = bk_key in SHARP_BOOKMAKERS
        for mkt in bk.get("markets", []):
            for o in mkt.get("outcomes", []):
                price = o.get("price", 0)
                if not price:
                    continue
                k = None
                if mkt["key"] == "h2h":
                    k = "1" if o["name"] == home else ("2" if o["name"] == away else "X")
                elif mkt["key"] == "totals":
                    k = f"{o['name']}_{o.get('point', '')}"
                elif mkt["key"] == "spreads" and o.get("point") == 0:
                    k = "DNB_1" if o["name"] == home else "DNB_2"
                elif mkt["key"] in ("btts", "bts") and o["name"].lower().startswith("yes"):
                    k = "BTTS"
                if k:
                    per_book[bk_key][k] = max(per_book[bk_key].get(k, 0.0), price)
                    if is_sharp:
                        sharp[k] = max(sharp.get(k, 0.0), price)
                    else:
                        soft[k] = max(soft.get(k, 0.0), price)
    for target in (sharp, soft):
        if {"1", "X"} <= target.keys():
            target["1X"] = round(1 / (1/target["1"] + 1/target["X"]), 3)
        if {"2", "X"} <= target.keys():
            target["X2"] = round(1 / (1/target["2"] + 1/target["X"]), 3)
        for side in ("1", "2"):
            if side in target and "Over_2.5" in target:
                target[f"{side}_Over"] = round(target[side] * target["Over_2.5"] * 0.85, 3)
    return sharp, soft, per_book

def get_true_market_probs(o1: float, ox: float, o2: float) -> Tuple[float, float, float]:
    if not (o1 and ox and o2):
        return 0.0, 0.0, 0.0
    inv1, invx, inv2 = 1/o1, 1/ox, 1/o2
    margin = (inv1 + invx + inv2) - 1.0
    if margin <= 0:
        return inv1, invx, inv2
    return inv1/(1+margin), invx/(1+margin), inv2/(1+margin)

# ══════════════════════════════════════════════════════════════════════════════
# 9. BETTING TOOLS
# ══════════════════════════════════════════════════════════════════════════════
def kelly_stake(stat_p: float, odds: float, bankroll: float, confidence: float = 1.0) -> float:
    b = odds - 1.0
    if b <= 0 or stat_p <= 0:
        return 0.0
    full_k = (stat_p * (b + 1) - 1) / b
    if full_k <= 0:
        return 0.0
    return round(min(full_k * KELLY_FRACTION * confidence, MAX_KELLY_PCT) * bankroll, 2)

def record_pick(conn: sqlite3.Connection, pick: dict, match_utc: datetime) -> str:
    uid = hashlib.md5(f"{pick['Match']}{pick['Market']}{pick['Date']}{pick['Time']}".encode()).hexdigest()[:12]
    conn.execute("""
        INSERT OR IGNORE INTO picks (id,run_ts,match_date,match_time,league,match,market,odds,stat_p,ev,kelly,pick_type,trend)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (uid, datetime.now(timezone.utc).isoformat(), pick["Date"], match_utc.isoformat(),
          pick.get("League", ""), pick["Match"], pick["Market"], pick["Odds"],
          pick.get("stat_p", 0), pick["EV"], pick.get("kelly_raw", 0),
          pick["Type"], pick["Trend"]))
    conn.commit()
    return uid

def get_trend(key: str, price: float, history: dict) -> Tuple[str, float]:
    prev = history.get(key)
    if prev is None or prev == 0:
        return "🆕 NEW", 0.0
    pct = (price - prev) / prev * 100
    if pct <= -3:
        return f"📉 {pct:+.1f}%", pct
    if pct >= +3:
        return f"📈 {pct:+.1f}%", pct
    return "↔ Stable", pct

# ══════════════════════════════════════════════════════════════════════════════
# 10. ARBITRAGE DETECTION
# ══════════════════════════════════════════════════════════════════════════════
def detect_arbitrage(per_book: dict, markets: list = ["1", "X", "2"]) -> Optional[dict]:
    if not FEATURES["arbitrage"]:
        return None
    best_odds: Dict[str, float] = {}
    best_book: Dict[str, str] = {}
    for bk, bk_odds in per_book.items():
        for m in markets:
            if m in bk_odds and bk_odds[m] > best_odds.get(m, 0):
                best_odds[m] = bk_odds[m]
                best_book[m] = bk
    if not all(m in best_odds for m in markets):
        return None
    inv_sum = sum(1.0 / best_odds[m] for m in markets)
    if inv_sum < 1.0:
        return {
            "margin_pct": round((1 - inv_sum) * 100, 3),
            "legs": {m: {"odds": best_odds[m], "book": best_book[m]} for m in markets},
        }
    return None

# ══════════════════════════════════════════════════════════════════════════════
# 11. EXCEL OUTPUT
# ══════════════════════════════════════════════════════════════════════════════
FILL = {"🎯 VALUE": PatternFill("solid", fgColor="C6EFCE"),
        "🔥 PATTERN": PatternFill("solid", fgColor="FFEB9C"),
        "⚠️ TRAP": PatternFill("solid", fgColor="FFC7CE"),
        "🔀 ARB": PatternFill("solid", fgColor="E2CFFF")}
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(color="FFFFFF", bold=True)

def write_excel(picks: List[dict], path: str, arbs: List[dict] = []) -> None:
    if not picks and not arbs:
        return
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        if picks:
            df_picks = pd.DataFrame(picks).sort_values(["Date", "Time", "EV"], ascending=[True, True, False]).reset_index(drop=True)
            summary = df_picks.groupby("Type").agg(Count=("EV","count"), Avg_EV=("EV","mean"), Best_EV=("EV","max")).reset_index().round(3)
            df_picks.to_excel(writer, index=False, sheet_name="Picks")
            for cell in writer.sheets["Picks"][1]:
                cell.fill, cell.font, cell.alignment = HDR_FILL, HDR_FONT, Alignment(horizontal="center")
            col_idx = list(df_picks.columns).index("Type")
            for row in writer.sheets["Picks"].iter_rows(min_row=2):
                fill = FILL.get(row[col_idx].value)
                if fill:
                    for cell in row:
                        cell.fill = fill
            for col in writer.sheets["Picks"].columns:
                writer.sheets["Picks"].column_dimensions[col[0].column_letter].width = max(len(str(c.value or "")) for c in col) + 4
            summary.to_excel(writer, index=False, sheet_name="Summary")
        if arbs:
            df_arb = pd.DataFrame(arbs)
            df_arb.to_excel(writer, index=False, sheet_name="Arbitrage")
            for cell in writer.sheets["Arbitrage"][1]:
                cell.fill, cell.font, cell.alignment = HDR_FILL, HDR_FONT, Alignment(horizontal="center")
            for col in writer.sheets["Arbitrage"].columns:
                writer.sheets["Arbitrage"].column_dimensions[col[0].column_letter].width = max(len(str(c.value or "")) for c in col) + 4

ICONS = {"VALUE": "🎯", "PATTERN": "🏆", "TRAP": "⚠️"}
TAGS  = {"VALUE": "🎯 VALUE", "PATTERN": "🔥 PATTERN", "TRAP": "⚠️ TRAP"}

def classify(stat_p: float, market_p: float, ev: float, sharp_trap: bool = False) -> Optional[str]:
    if sharp_trap:
        return "TRAP"
    stat_pct, market_pct = stat_p * 100, market_p * 100
    if ev >= EV_THRESHOLD:
        return "VALUE"
    if stat_pct >= PATTERN_THRESHOLD:
        return "PATTERN"
    if (market_pct - stat_pct) >= TRAP_MARGIN:
        return "TRAP"
    return None

# ══════════════════════════════════════════════════════════════════════════════
# 12. MAIN SCAN (V36 ENHANCED)
# ══════════════════════════════════════════════════════════════════════════════
def run_scan() -> None:
    log.info("=" * 80)
    log.info("🚀  ORACLE V36.0 ENTERPRISE — Merged V35.2 + Risk Management")
    log.info("=" * 80)

    conn = get_db()
    all_picks = []
    all_arbs = []
    history = {row["key"]: row["price"] for row in conn.execute("SELECT key, price FROM odds_history").fetchall()}
    now_utc = datetime.now(timezone.utc).date()
    target_dates = {now_utc + timedelta(days=d) for d in range(LOOKAHEAD_DAYS)}

    for league in LEAGUE_KEYS:
        log.info(f"\n📡  {league}")
        try:
            resp = requests.get(f"https://api.the-odds-api.com/v4/sports/{league}/odds/",
                                params={"apiKey": ODDS_API_KEY, "regions": "eu", "markets": "h2h,totals,spreads"},
                                timeout=15).json()
        except requests.RequestException:
            continue
        if not isinstance(resp, list):
            continue

        for match in resp:
            try:
                m_utc = datetime.strptime(match["commence_time"], "%Y-%m-%dT%H:%M:%SZ").replace(tzinfo=timezone.utc)
            except (KeyError, ValueError):
                continue
            if m_utc.date() not in target_dates:
                continue

            local = m_utc + timedelta(hours=UTC_OFFSET_HOURS)
            home, away = match["home_team"], match["away_team"]

            sharp_odds, soft_odds, per_book = parse_sharp_and_soft_odds(match, home, away)
            if not sharp_odds.get("1"):
                sharp_odds = soft_odds
            if "1" not in soft_odds:
                continue
            if soft_odds.get("1", 0) < 1.15 and soft_odds.get("2", 0) < 1.15:
                continue

            arb = detect_arbitrage(per_book)
            if arb:
                all_arbs.append({
                    "Date": local.strftime("%d/%m"), "Time": local.strftime("%H:%M"),
                    "Match": f"{home} – {away}",
                    "League": league.replace("soccer_","").replace("_"," ").title(),
                    "Margin %": f"{arb['margin_pct']:.3f}%",
                    "Legs": " | ".join(f"{m}: {v['odds']} @ {v['book']}" for m, v in arb["legs"].items()),
                })
                log.info(f"      🔀 ARB {arb['margin_pct']:.3f}% — {home} vs {away}")

            _h_row = conn.execute("SELECT team_id FROM team_cache WHERE team_name=?", (home,)).fetchone()
            _a_row = conn.execute("SELECT team_id FROM team_cache WHERE team_name=?", (away,)).fetchone()
            if _h_row and _a_row:
                _ep1, _epx, _ep2 = elo_probabilities(get_elo(conn, _h_row["team_id"]),
                                                     get_elo(conn, _a_row["team_id"]))
                if not any((p * soft_odds.get(m, 0)) >= PRE_FILTER_EV for p, m in zip([_ep1, _epx, _ep2], ["1", "X", "2"])):
                    continue

            ref_name = match.get("fixture", {}).get("referee") or match.get("referee", "")

            for _m in MARKET_LABELS:
                _ov = soft_odds.get(_m)
                if _ov:
                    _key = f"{home}|{away}|{_m}"
                    if _key not in history:
                        history[_key] = _ov

            # V36: Pass match_utc for fatigue calculation
            stats = get_full_analysis(conn, home, away, league, ref_name, m_utc)
            if not stats:
                continue

            flagged = False
            true_p1, true_px, true_p2 = get_true_market_probs(sharp_odds.get("1", 0),
                                                              sharp_odds.get("X", 0),
                                                              sharp_odds.get("2", 0))
            true_probs = {"1": true_p1, "X": true_px, "2": true_p2}
            w_dc, w_elo = get_dynamic_weights(stats.get("_matches_played", 0))

            # V36: Hedge note preparation
            hedge_note = ""
            if FEATURES["hedging"] and ("1" in soft_odds and "X" in soft_odds and "2" in soft_odds):
                # Θα υπολογιστεί ανά market
                pass

            for market in MARKET_LABELS:
                sp = stats.get(market)
                ov = soft_odds.get(market)
                if sp is None or ov is None or ov <= 1.0:
                    continue

                # V36: Auto-calibration adjustment
                cal_mult = get_brier_calibration(market)
                sp_cal = sp / cal_mult  # Μείωση πιθανότητας αν το μοντέλο υπερεκτιμά
                sp_cal = max(0.05, min(0.95, sp_cal))

                mp = true_probs.get(market) if market in true_probs else (1.0 / sharp_odds.get(market, ov))

                if market in ("1", "X", "2"):
                    _mk = {"1": ("_elo_p1","_pyth_p1","_form_p1"),
                           "X": ("_elo_px","_pyth_px","_form_px"),
                           "2": ("_elo_p2","_pyth_p2","_form_p2")}[market]
                    elo_p_val  = stats.get(_mk[0]) or sp_cal
                    pyth_p_val = stats.get(_mk[1]) or sp_cal
                    form_p_val = stats.get(_mk[2]) or sp_cal
                    sp_final, confidence, agreement = get_ensemble_stat_p(
                        sp_cal, elo_p_val, mp, pyth_p_val, form_p_val, w_dc, w_elo
                    )
                else:
                    sp_final = sp_cal
                    confidence, agreement = 0.99, "HIGH"

                # V36: Hedging calculation
                hedge_note = ""
                if FEATURES["hedging"] and market in ('1', '2'):
                    odds_1 = soft_odds.get('1', 0)
                    odds_x = soft_odds.get('X', 0)
                    odds_2 = soft_odds.get('2', 0)
                    if odds_1 and odds_x and odds_2:
                        hedge_note = calculate_hedge_note(market, odds_1, odds_x, odds_2)

                key = f"{home}|{away}|{market}"
                trend_label, trend_pct = get_trend(key, ov, history)
                history[key] = ov

                sharp_trap = False
                if market == "1":
                    _, opp_pct = get_trend(f"{home}|{away}|2", soft_odds.get("2", 0), history)
                    if opp_pct <= -5.0:
                        sharp_trap = True
                elif market == "2":
                    _, opp_pct = get_trend(f"{home}|{away}|1", soft_odds.get("1", 0), history)
                    if opp_pct <= -5.0:
                        sharp_trap = True

                ev = round(sp_final * ov, 4)
                typ = classify(sp_final, mp, ev, sharp_trap)
                if typ is None:
                    continue

                if not flagged:
                    log.info(f"\n      🏟️  {home} – {away}  {local.strftime('%d/%m %H:%M')}")
                    log.info(f"         ⭐ Specials | {stats['Specials_Corners']} | {stats['Specials_Cards']} | Scorer: {stats['Specials_Scorer']}")
                    flagged = True

                sp_pct, mp_pct, conf_pct = int(sp_final * 100), int(mp * 100), int(confidence * 100)
                kelly = kelly_stake(sp_final, ov, BANKROLL, confidence)

                log.info(f"         {ICONS[typ]} {TAGS[typ]:<13} | {market:<10} | Odds: {ov:<6} | Stat: {sp_pct}% | Sharp: {mp_pct}% | EV: {ev}  Kelly: €{kelly} | ha={stats.get('_ha_dynamic',1.28):.2f} h_ppg={stats.get('_h_momentum',0.5):.2f}")

                pick = {
                    "Date": local.strftime("%d/%m"), "Time": local.strftime("%H:%M"),
                    "League": league.replace("soccer_","").replace("_"," ").title(),
                    "Match": f"{home} – {away}",
                    "Market": market, "Odds": ov, "Book %": f"{mp_pct}%", "Stat %": f"{sp_pct}%",
                    "stat_p": sp_final, "EV": ev, "Kelly": f"€{kelly}", "kelly_raw": kelly,
                    "Trend": trend_label, "Type": TAGS[typ], "Confidence": f"{conf_pct}%",
                    "Agreement": agreement, "System": "V36-Enterprise",
                    "λ": round(stats.get("_lam", 0), 3),
                    "μ": round(stats.get("_mu", 0), 3),
                    "Home Adv": round(stats.get("_ha_dynamic", 1.28), 3),
                    "H PPG": round(stats.get("_h_momentum", 0.5), 3),
                    "A PPG": round(stats.get("_a_momentum", 0.5), 3),
                    "Proj Corners": stats["Specials_Corners"],
                    "Proj Cards": stats["Specials_Cards"],
                    "Top Scorer Pick": stats["Specials_Scorer"],
                    "Hedge Note": hedge_note,  # V36
                }
                all_picks.append(pick)
                record_pick(conn, pick, m_utc)

    now = datetime.now(timezone.utc).isoformat()
    conn.executemany("INSERT OR REPLACE INTO odds_history (key, price, updated) VALUES (?,?,?)",
                     [(k, v, now) for k, v in history.items()])
    conn.commit()
    write_excel(all_picks, OUTPUT_FILE, all_arbs)
    log.info(f"\n📂  Saved → {OUTPUT_FILE}  ({len(all_picks)} picks, {len(all_arbs)} arbs)")
    conn.close()

if __name__ == "__main__":
    run_scan()
