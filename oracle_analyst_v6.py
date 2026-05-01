"""
╔══════════════════════════════════════════════════════════════════════════════╗
║  ORACLE ANALYST V6 — ENHANCED PICK SELECTOR (Dynamic Threshold Bonus)      ║
║  • Copies λ, μ, Home Adv, H PPG, A PPG to Top Picks sheet                  ║
║  • Empirical confidence from historical ledger                              ║
║  • Auto‑calibrated type bonuses                                             ║
║  • Correlation penalty                                                      ║
║  • Model Strength Score (synergy of internal model indicators)              ║
║  • NEW: Threshold Bonus (learns winning conditions from Ledger)             ║
║  • NEW: Threshold Details column in Score Breakdown                         ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

import sys
import os
import glob
import math
import pandas as pd
import numpy as np
from itertools import combinations
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# =====================================================================
# CONFIGURATION – weights updated for V6
# =====================================================================
MIN_PICKS = 10
MAX_PICKS = 15
HIGH_ODDS_THRESHOLD = 3.00

# Weights now include model strength (6%)
WEIGHT_EV          = 0.33
WEIGHT_CONFIDENCE  = 0.23
WEIGHT_STAT_P      = 0.18
WEIGHT_ODDS        = 0.10
WEIGHT_TREND       = 0.10
WEIGHT_MODEL       = 0.06   # νέο βάρος για λ/μ/HomeAdv/H PPG/A PPG

TYPE_BONUS = {
    "🎯 VALUE":   0.15,
    "🔥 PATTERN": 0.05,
    "⚠️ TRAP":   -0.10,
}

COVERAGE_MAP = {
    "1": ["1X", "DNB_1", "1_Over", "X"],
    "2": ["X2", "DNB_2", "2_Over", "X"],
    "X": ["1X", "X2"],
    "Over_2.5": ["BTTS", "1_Over", "2_Over"],
    "BTTS": ["Over_2.5"],
    "DNB_1": ["1", "1X"],
    "DNB_2": ["2", "X2"],
}

ODDS_SWEET_SPOT = 2.00
ODDS_SIGMA = 0.80

OUTPUT_FILE = "Oracle_Analyst_Report_v6.xlsx"
LEDGER_FILE = "Oracle_Historical_Ledger.xlsx"

C_GRAY_MID  = "D9D9D9"
C_NAVY_MID  = "2E75B6"

CORRELATION_PENALTY_THRESHOLD = 0.65
CORRELATION_PENALTY_AMOUNT = 0.10

# =====================================================================
# 1. LOAD HISTORICAL LEDGER & COMPUTE WIN RATES
# =====================================================================
def load_ledger():
    if os.path.exists(LEDGER_FILE):
        try:
            df = pd.read_excel(LEDGER_FILE, sheet_name="Ledger")
            df = df[df["Result"].isin(["WIN", "LOSS"])].copy()
            df["Win"] = (df["Result"] == "WIN").astype(int)
            return df
        except Exception as e:
            print(f"⚠️ Could not load ledger: {e}")
    return pd.DataFrame()

def compute_empirical_rates(ledger):
    if ledger.empty:
        return {}, {}, {}, {}
    market_rates = ledger.groupby("Market")["Win"].agg(['mean', 'count']).to_dict('index')
    ledger['Type_clean'] = ledger['Type'].str.extract(r'(VALUE|PATTERN|TRAP)').fillna('OTHER')
    type_rates = ledger.groupby("Type_clean")["Win"].mean().to_dict()
    league_rates = ledger.groupby("League")["Win"].mean().to_dict()
    def odds_range(odds):
        if odds < 1.5: return "1.00-1.50"
        elif odds < 2.0: return "1.50-2.00"
        elif odds < 2.5: return "2.00-2.50"
        elif odds < 3.0: return "2.50-3.00"
        else: return "3.00+"
    ledger['OddsRange'] = ledger['Odds'].apply(odds_range)
    odds_rates = ledger.groupby("OddsRange")["Win"].mean().to_dict()
    return market_rates, type_rates, league_rates, odds_rates

def adjust_weights_from_ledger(ledger):
    if ledger.empty:
        return WEIGHT_EV, WEIGHT_CONFIDENCE, WEIGHT_STAT_P, WEIGHT_ODDS, WEIGHT_TREND, WEIGHT_MODEL, TYPE_BONUS
    for t in TYPE_BONUS:
        if t == "🎯 VALUE":
            win_rate = ledger[ledger['Type_clean']=='VALUE']['Win'].mean() if 'VALUE' in ledger['Type_clean'].values else 0.5
            TYPE_BONUS[t] = (win_rate - 0.5) * 0.4
        elif t == "🔥 PATTERN":
            win_rate = ledger[ledger['Type_clean']=='PATTERN']['Win'].mean() if 'PATTERN' in ledger['Type_clean'].values else 0.5
            TYPE_BONUS[t] = (win_rate - 0.5) * 0.3
        elif t == "⚠️ TRAP":
            win_rate = ledger[ledger['Type_clean']=='TRAP']['Win'].mean() if 'TRAP' in ledger['Type_clean'].values else 0.5
            TYPE_BONUS[t] = (win_rate - 0.5) * 0.5
    return WEIGHT_EV, WEIGHT_CONFIDENCE, WEIGHT_STAT_P, WEIGHT_ODDS, WEIGHT_TREND, WEIGHT_MODEL, TYPE_BONUS

# =====================================================================
# 2. LOAD ORACLE EXCEL (V36 output)
# =====================================================================
def find_oracle_excel(path: str = None) -> str:
    if path and os.path.exists(path):
        return path
    candidates = sorted(glob.glob("Oracle_V*.xlsx") + glob.glob("Ultimate_Apex*.xlsx"),
                        key=os.path.getmtime, reverse=True)
    if candidates:
        print(f"📂  Found: {candidates[0]}")
        return candidates[0]
    raise FileNotFoundError("No Oracle Excel found.")

def load_picks(excel_path: str) -> pd.DataFrame:
    xl = pd.read_excel(excel_path, sheet_name=None)
    if "Picks" not in xl:
        raise ValueError(f"Sheet 'Picks' not found in {excel_path}")
    df = xl["Picks"].copy()
    print(f"✅ Loaded {len(df)} picks from {excel_path}")
    
    df["EV"] = pd.to_numeric(df["EV"], errors="coerce").fillna(1.0)
    if "Confidence" in df.columns:
        df["Confidence_num"] = df["Confidence"].astype(str).str.rstrip("%").apply(
            lambda x: float(x)/100 if x.replace(".","").isdigit() else 0.5
        )
    else:
        df["Confidence_num"] = 0.5
    if "Stat %" in df.columns:
        df["StatP_num"] = df["Stat %"].astype(str).str.rstrip("%").apply(
            lambda x: float(x)/100 if x.replace(".","").isdigit() else 0.5
        )
    else:
        df["StatP_num"] = 0.5
    df["Odds"] = pd.to_numeric(df["Odds"], errors="coerce").fillna(1.5)
    df["Trend_score"] = df.get("Trend", "").apply(lambda t: 0.8 if "📈" in str(t) else 0.6)
    if "Agreement" in df.columns:
        df["Agreement_num"] = df["Agreement"].map({"HIGH":1.0, "MEDIUM":0.65, "LOW":0.30}).fillna(0.5)
    else:
        df["Agreement_num"] = df["Confidence_num"]
    # V36 specific columns: λ, μ, Home Adv, H PPG, A PPG
    for col in ["λ", "μ", "Home Adv", "H PPG", "A PPG"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0 if col in ["λ","μ"] else 1.28 if col=="Home Adv" else 0.5)
    return df

# =====================================================================
# 3. EMPIRICAL CONFIDENCE (same as V5)
# =====================================================================
def get_empirical_confidence(row, market_rates, type_rates, league_rates, odds_rates):
    market = row.get("Market", "")
    pick_type = str(row.get("Type", ""))
    league = row.get("League", "")
    odds = row.get("Odds", 2.0)
    if odds < 1.5: odds_range = "1.00-1.50"
    elif odds < 2.0: odds_range = "1.50-2.00"
    elif odds < 2.5: odds_range = "2.00-2.50"
    elif odds < 3.0: odds_range = "2.50-3.00"
    else: odds_range = "3.00+"
    m_rate = market_rates.get(market, {}).get('mean', 0.5) if market_rates else 0.5
    if "VALUE" in pick_type: t_key = "VALUE"
    elif "PATTERN" in pick_type: t_key = "PATTERN"
    elif "TRAP" in pick_type: t_key = "TRAP"
    else: t_key = "OTHER"
    t_rate = type_rates.get(t_key, 0.5) if type_rates else 0.5
    l_rate = league_rates.get(league, 0.5) if league_rates else 0.5
    o_rate = odds_rates.get(odds_range, 0.5) if odds_rates else 0.5
    confidence = (m_rate * 0.3 + t_rate * 0.3 + l_rate * 0.2 + o_rate * 0.2)
    return min(1.0, max(0.0, confidence))

# =====================================================================
# 4. MODEL STRENGTH SCORE (NEW IN V6)
# =====================================================================
def compute_model_strength(row):
    try:
        lam = float(row.get("λ", 1.5))
        mu  = float(row.get("μ", 1.5))
        ha  = float(row.get("Home Adv", 1.28))
        h_ppg = float(row.get("H PPG", 0.5))
        a_ppg = float(row.get("A PPG", 0.5))
    except (ValueError, TypeError):
        return 0.5
    
    market = str(row.get("Market", ""))
    goal_diff = lam - mu
    form_diff = h_ppg - a_ppg
    
    if market in ("1", "1X", "DNB_1", "1_Over"):
        raw = goal_diff * 0.35 + (ha - 1.0) * 0.35 + form_diff * 0.30
    elif market in ("2", "X2", "DNB_2", "2_Over"):
        raw = -goal_diff * 0.35 - (ha - 1.0) * 0.35 - form_diff * 0.30
    elif market == "Over_2.5":
        raw = (lam + mu - 2.5) * 0.50 + (ha - 1.0) * 0.20 + (h_ppg + a_ppg - 1.0) * 0.30
    elif market == "Under_2.5":
        raw = (2.5 - lam - mu) * 0.50 - (ha - 1.0) * 0.20 - (h_ppg + a_ppg - 1.0) * 0.30
    elif market == "X":
        raw = 0.5 - abs(goal_diff) * 0.30 - abs(form_diff) * 0.20
    else:
        return 0.5
    
    score = 1.0 / (1.0 + math.exp(-raw * 3))
    return round(score, 3)

# =====================================================================
# 5. DYNAMIC THRESHOLD ANALYSIS (NEW)
# =====================================================================
def find_optimal_thresholds(ledger, min_samples=20):
    """
    Βρίσκει τα βέλτιστα κατώφλια για κάθε αγορά και κάθε δείκτη.
    Επιστρέφει dict: {market: {feature: (threshold, direction)}}
    """
    if ledger.empty:
        return {}

    features = [c for c in ["λ (Lambda)", "μ (Mu)", "Home_Adv", "H_PPG", "A_PPG",
                             "Home_xG", "Away_xG", "Total_xG"]
                if c in ledger.columns]

    markets = ledger["Market"].unique() if "Market" in ledger.columns else []

    optimal = {}
    for market in markets:
        market_df = ledger[ledger["Market"] == market]
        if len(market_df) < min_samples:
            continue
        optimal[market] = {}
        for feat in features:
            feat_vals = market_df[feat].dropna()
            if len(feat_vals) < min_samples:
                continue

            thresholds = np.percentile(feat_vals, np.linspace(10, 90, 20))
            best_delta = -np.inf
            best_thresh = None
            best_direction = None

            for thresh in thresholds:
                above = market_df[market_df[feat] >= thresh]
                below = market_df[market_df[feat] < thresh]
                if len(above) < min_samples/2 or len(below) < min_samples/2:
                    continue

                if "PnL" in market_df.columns:
                    metric_above = above["PnL"].mean()
                    metric_below = below["PnL"].mean()
                else:
                    metric_above = above["Win"].mean()
                    metric_below = below["Win"].mean()

                delta = metric_above - metric_below
                if abs(delta) > abs(best_delta):
                    best_delta = delta
                    best_thresh = thresh
                    best_direction = "above" if delta > 0 else "below"

            if best_thresh is not None and abs(best_delta) > 0.01:
                optimal[market][feat] = (best_thresh, best_direction)

    return optimal

def format_threshold_criteria(market, optimal_thresholds, row=None):
    """
    Επιστρέφει ένα string που περιγράφει τα κατώφλια για την αγορά,
    και αν δοθεί row, ποια από αυτά ικανοποιούνται.
    """
    if not optimal_thresholds or market not in optimal_thresholds:
        return "No thresholds"

    criteria = optimal_thresholds[market]
    parts = []
    for feat, (thresh, direction) in criteria.items():
        # Συντομογραφία του δείκτη
        short_name = feat.replace(" (Lambda)", "").replace(" (Mu)", "").replace("Home_Adv", "HA").replace("H_PPG", "HPPG").replace("A_PPG", "APPG")
        symbol = "≥" if direction == "above" else "<"
        part = f"{short_name} {symbol} {thresh:.2f}"
        
        # Αν έχουμε και row, ελέγχουμε αν ικανοποιείται
        if row is not None:
            val = None
            if feat == "λ (Lambda)":
                val = row.get("λ")
            elif feat == "μ (Mu)":
                val = row.get("μ")
            elif feat == "Home_Adv":
                val = row.get("Home Adv")
            elif feat == "H_PPG":
                val = row.get("H PPG")
            elif feat == "A_PPG":
                val = row.get("A PPG")
            
            if val is not None:
                if direction == "above" and val >= thresh:
                    part += " ✅"
                elif direction == "below" and val < thresh:
                    part += " ✅"
                else:
                    part += " ❌"
        parts.append(part)
    return ", ".join(parts) if parts else "None"

def compute_threshold_bonus(row, optimal_thresholds, bonus_amount=0.05):
    """
    Επιστρέφει (bonus, details_string) όπου details_string εξηγεί τι ενεργοποιήθηκε.
    """
    if not optimal_thresholds:
        return 0.0, ""

    market = row.get("Market", "")
    if market not in optimal_thresholds:
        return 0.0, ""

    features_used = []
    details = []
    for feat, (thresh, direction) in optimal_thresholds[market].items():
        val = None
        short_name = feat.replace(" (Lambda)", "λ").replace(" (Mu)", "μ").replace("Home_Adv", "HA").replace("H_PPG", "HPPG").replace("A_PPG", "APPG")
        if feat == "λ (Lambda)":
            val = row.get("λ", None)
        elif feat == "μ (Mu)":
            val = row.get("μ", None)
        elif feat == "Home_Adv":
            val = row.get("Home Adv", None)
        elif feat == "H_PPG":
            val = row.get("H PPG", None)
        elif feat == "A_PPG":
            val = row.get("A PPG", None)

        if val is None:
            continue

        satisfied = False
        if direction == "above" and val >= thresh:
            satisfied = True
        elif direction == "below" and val < thresh:
            satisfied = True

        if satisfied:
            features_used.append(short_name)
            details.append(f"{short_name} {direction} {thresh:.2f} (val={val:.3f})")

    bonus = min(bonus_amount, len(features_used) * 0.02)
    detail_str = "; ".join(details) if details else "None"
    return bonus, detail_str

# =====================================================================
# 6. COMPOSITE SCORE (updated for V6 + Threshold Bonus)
# =====================================================================
def odds_sweet_spot_score(odds: float) -> float:
    return math.exp(-0.5 * ((odds - ODDS_SWEET_SPOT) / ODDS_SIGMA) ** 2)

def compute_composite_score(row, market_rates, type_rates, league_rates, odds_rates, weights, optimal_thresholds):
    ev_score = min(1.0, (row["EV"] - 1.0) / 0.15)
    emp_conf = get_empirical_confidence(row, market_rates, type_rates, league_rates, odds_rates)
    conf_score = emp_conf * row.get("Agreement_num", 0.5)
    stat_score = max(0.0, min(1.0, (row["StatP_num"] - 0.40) / 0.40))
    odds_score = odds_sweet_spot_score(row["Odds"])
    trend_score = row["Trend_score"]
    model_score = compute_model_strength(row)

    score = (ev_score * weights[0] + conf_score * weights[1] +
             stat_score * weights[2] + odds_score * weights[3] +
             trend_score * weights[4] + model_score * weights[5])

    # --- Threshold Bonus ---
    threshold_bonus, threshold_detail = compute_threshold_bonus(row, optimal_thresholds)
    score += threshold_bonus

    pick_type = str(row.get("Type", ""))
    for tag, bonus in TYPE_BONUS.items():
        if tag in pick_type:
            score += bonus
            break
    return round(max(0.0, min(1.0, score)), 4), threshold_bonus, threshold_detail

# =====================================================================
# 7. CORRELATION PENALTY
# =====================================================================
def compute_market_correlation(ledger):
    if ledger.empty:
        return {}
    ledger['MatchKey'] = ledger['Match'] + "|" + ledger['Date'].astype(str)
    pivot = ledger.pivot_table(index='MatchKey', columns='Market', values='Win', aggfunc='first')
    corr = pivot.corr(min_periods=3)
    corr_dict = {}
    for m1 in corr.columns:
        for m2 in corr.columns:
            if m1 < m2 and not pd.isna(corr.loc[m1, m2]):
                corr_dict[(m1, m2)] = corr.loc[m1, m2]
    return corr_dict

def apply_correlation_penalty(selected, all_picks, corr_dict):
    if not corr_dict:
        return selected
    selected['Penalty'] = 0.0
    for match in selected['Match'].unique():
        match_picks = selected[selected['Match'] == match]
        if len(match_picks) < 2:
            continue
        markets = match_picks['Market'].tolist()
        for m1, m2 in combinations(markets, 2):
            corr = corr_dict.get((m1, m2), corr_dict.get((m2, m1), 0))
            if abs(corr) >= CORRELATION_PENALTY_THRESHOLD:
                selected.loc[match_picks.index, 'Penalty'] += CORRELATION_PENALTY_AMOUNT / 2
    selected['Composite_Score'] = selected['Composite_Score'] - selected['Penalty']
    selected['Composite_Score'] = selected['Composite_Score'].clip(0, 1)
    return selected

# =====================================================================
# 8. SELECT TOP PICKS
# =====================================================================
def select_top_picks(df, market_rates, type_rates, league_rates, odds_rates, weights, corr_dict, optimal_thresholds):
    df = df.copy()
    scores_details = df.apply(
        lambda r: compute_composite_score(r, market_rates, type_rates, league_rates, odds_rates, weights, optimal_thresholds),
        axis=1
    )
    df['Composite_Score'] = [x[0] for x in scores_details]
    df['Threshold_Bonus'] = [x[1] for x in scores_details]
    df['Threshold_Detail'] = [x[2] for x in scores_details]
    
    df = df.sort_values('Composite_Score', ascending=False).reset_index(drop=True)
    
    selected = []
    match_count = {}
    for _, row in df.iterrows():
        match_key = str(row.get("Match", ""))
        if match_count.get(match_key, 0) >= 2:
            continue
        selected.append(row)
        match_count[match_key] = match_count.get(match_key, 0) + 1
        if len(selected) >= MAX_PICKS:
            break
    
    result = pd.DataFrame(selected).reset_index(drop=True)
    result = apply_correlation_penalty(result, df, corr_dict)
    result = result.sort_values('Composite_Score', ascending=False).reset_index(drop=True)
    
    if len(result) < MIN_PICKS:
        for _, row in df.iterrows():
            if not any((r['Match'] == row['Match'] and r['Market'] == row['Market']) for r in selected):
                selected.append(row)
                if len(selected) >= MIN_PICKS:
                    break
        result = pd.DataFrame(selected).reset_index(drop=True)
    
    print(f"✅ Selected {len(result)} picks (from {len(df)} total)")
    return result

# =====================================================================
# 9. COVERAGE & HEDGE HELPERS
# =====================================================================
def find_coverage(pick_row, all_picks):
    match  = str(pick_row.get("Match",  ""))
    market = str(pick_row.get("Market", ""))
    same   = all_picks[all_picks["Match"] == match]
    results = []
    for cov_mkt in COVERAGE_MAP.get(market, []):
        rows = same[same["Market"] == cov_mkt]
        if not rows.empty:
            cov = rows.iloc[0]
            cov_odds = float(cov.get("Odds", 0))
            if cov_odds > 1.0:
                results.append({"coverage_market": cov_mkt, "coverage_odds": cov_odds, "source": "excel"})
    if not results:
        for cov_mkt in COVERAGE_MAP.get(market, [])[:2]:
            results.append({"coverage_market": cov_mkt, "coverage_odds": None, "source": "suggested"})
    return results

def calc_hedge(main_odds, cover_odds, main_stake):
    if not (cover_odds and cover_odds > 1.0 and main_odds > 1.0):
        return {}
    cover_stake  = round(main_stake * main_odds / cover_odds, 2)
    profit_main  = round(main_stake * (main_odds - 1) - cover_stake, 2)
    profit_cover = round(cover_stake * (cover_odds - 1) - main_stake, 2)
    return {
        "cover_stake": cover_stake,
        "profit_main": profit_main, "profit_cover": profit_cover,
        "is_arb": profit_main > 0 and profit_cover > 0,
    }

# =====================================================================
# 10. EXCEL STYLE HELPERS & WRITER
# =====================================================================
C_NAVY      = "1F4E79"; C_NAVY_MID  = "2E75B6"; C_NAVY_LT = "D6E4F0"
C_WHITE     = "FFFFFF"; C_GRAY_LT   = "F2F2F2"; C_GRAY_MID = "D9D9D9"
C_GREEN_BG  = "E2EFDA"; C_GREEN_DK  = "375623"; C_GREEN_CELL = "C6EFCE"
C_AMBER_BG  = "FFEB9C"; C_AMBER_DK  = "7A4F00"
C_RED_BG    = "FFC7CE"; C_RED_DK    = "9C0006"
C_ORANGE_LT = "FCE4D6"; C_ORANGE    = "C55A11"
C_COVER_BG  = "EBF4EC"; C_ARB_BG    = "D5F0D5"
C_PATTERN   = "FFF2CC"; C_BLACK     = "000000"

_TH  = Side(border_style="thin",   color="BFBFBF")
_MED = Side(border_style="medium", color=C_NAVY)
_B   = Border(left=_TH, right=_TH, top=_TH,   bottom=_TH)
_BH  = Border(left=_TH, right=_TH, top=_MED,  bottom=_MED)

def _h(ws, r, c, v, bg=C_NAVY, fg=C_WHITE, sz=9, bold=True):
    cell = ws.cell(row=r, column=c, value=v)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(color=fg, bold=bold, size=sz, name="Calibri")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _BH
    return cell

def _c(ws, r, c, v, bg=None, fg=C_BLACK, bold=False, align="center", sz=9, wrap=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font      = Font(color=fg, bold=bold, size=sz, name="Calibri")
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border    = _B
    if bg: cell.fill = PatternFill("solid", fgColor=bg)
    return cell

def _type_bg(t):
    if "VALUE"   in str(t): return C_GREEN_CELL
    if "PATTERN" in str(t): return C_PATTERN
    if "TRAP"    in str(t): return C_RED_BG
    return C_WHITE

def _score_fg(s):
    if s >= 0.70: return C_GREEN_DK
    if s >= 0.45: return C_AMBER_DK
    return C_RED_DK

def _score_bg(s):
    if s >= 0.70: return C_GREEN_BG
    if s >= 0.45: return C_AMBER_BG
    return C_RED_BG

def _alt(i): return C_GRAY_LT if i % 2 == 0 else C_WHITE

def _grade(s):
    if s >= 0.75: return "S"
    if s >= 0.60: return "A"
    if s >= 0.45: return "B"
    return "C"

def _set_tab(ws, color): ws.sheet_properties.tabColor = color

def write_report(selected: pd.DataFrame, all_picks: pd.DataFrame, path: str, optimal_thresholds: dict) -> None:
    from datetime import datetime
    wb = Workbook()

    # Sheet 1: TOP PICKS
    ws1 = wb.active
    ws1.title = "Top Picks"
    _set_tab(ws1, C_NAVY)
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "C5"
    for r, h in {1: 36, 2: 20, 3: 16, 4: 26}.items():
        ws1.row_dimensions[r].height = h

    n_cols = 23
    ws1.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    c = ws1["A1"]
    c.value     = f"  ◆  ORACLE ANALYST V6  ◆  Top {len(selected)} Picks  ·  {datetime.now().strftime('%d %b %Y  %H:%M')}"
    c.font      = Font(bold=True, size=15, color=C_WHITE, name="Calibri")
    c.fill      = PatternFill("solid", fgColor=C_NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")

    n_val  = selected["Type"].str.contains("VALUE",   na=False).sum()
    n_pat  = selected["Type"].str.contains("PATTERN", na=False).sum()
    n_trap = selected["Type"].str.contains("TRAP",    na=False).sum()
    n_high = (selected["Odds"] >= HIGH_ODDS_THRESHOLD).sum()
    avg_ev = selected["EV"].mean()
    avg_sc = selected["Composite_Score"].mean()

    kpi_data = [
        (f"🎯  VALUE: {n_val}",       "A", "C", C_GREEN_CELL),
        (f"🔥  PATTERN: {n_pat}",     "D", "F", C_PATTERN),
        (f"⚠️  TRAP: {n_trap}",        "G", "I", C_RED_BG),
        (f"🔶  High Odds: {n_high}",  "J", "L", C_ORANGE_LT),
        (f"📊  Avg EV: {avg_ev:.3f}", "M", "O", C_NAVY_LT),
        (f"⭐  Avg Score: {avg_sc:.3f}", "P", get_column_letter(n_cols), C_NAVY_LT),
    ]
    for text, start, end, bg in kpi_data:
        ws1.merge_cells(f"{start}2:{end}2")
        kc = ws1[f"{start}2"]
        kc.value     = text
        kc.font      = Font(bold=True, size=9, color=C_NAVY, name="Calibri")
        kc.fill      = PatternFill("solid", fgColor=bg)
        kc.alignment = Alignment(horizontal="center", vertical="center")
        kc.border    = _B

    ws1.merge_cells(f"A3:{get_column_letter(n_cols)}3")
    s3 = ws1["A3"]
    s3.value     = "   SELECTED PICKS  ·  Ranked by Composite Score (incl. Model Strength & Threshold Bonus)  ·  Max 2 picks per match  ·  Correlation-penalised"
    s3.font      = Font(size=8, color="595959", name="Calibri")
    s3.fill      = PatternFill("solid", fgColor=C_NAVY_LT)
    s3.alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        "#", "Date", "Time", "League", "Match", "Market", "Odds", "Type",
        "Stat %", "Sharp %", "EV",
        "λ", "μ", "Home Adv", "H PPG", "A PPG",
        "Confidence", "Agreement", "Trend",
        "Score", "Grade", "Coverage", "Hedge Note"
    ]
    col_w = [
        4, 7, 6, 17, 30, 10, 7, 14,
        8, 8, 7,
        7, 7, 8, 7, 7,
        11, 10, 14,
        8, 7, 42, 34
    ]
    for col, (hdr, w) in enumerate(zip(headers, col_w), 1):
        _h(ws1, 4, col, hdr, sz=9)
        ws1.column_dimensions[get_column_letter(col)].width = w

    for rank, (_, row) in enumerate(selected.iterrows(), 1):
        r      = rank + 4
        ws1.row_dimensions[r].height = 22
        score  = float(row.get("Composite_Score", 0))
        g      = _grade(score)
        is_hi  = row["Odds"] >= HIGH_ODDS_THRESHOLD
        row_bg = C_ORANGE_LT if is_hi else _type_bg(row.get("Type",""))
        row_bg = row_bg if row_bg != C_WHITE else _alt(rank)

        coverages = find_coverage(row, all_picks) if is_hi else []
        cov_parts = []
        for cv in coverages:
            sym = "✅" if cv["source"] == "excel" else "💡"
            odds_str = f"@ {cv['coverage_odds']}" if cv["coverage_odds"] else "(find odds)"
            cov_parts.append(f"{sym} {cv['coverage_market']} {odds_str}")
        cov_txt = "  |  ".join(cov_parts) if cov_parts else "—"

        hedge_txt = ""
        for cv in coverages:
            if cv["coverage_odds"] and isinstance(cv["coverage_odds"], float):
                h = calc_hedge(row["Odds"], cv["coverage_odds"], float(row.get("kelly_raw", 10.0) or 10.0))
                if h:
                    arb = "  ✅ ARB" if h["is_arb"] else ""
                    hedge_txt = f"{cv['coverage_market']}: cover €{h['cover_stake']}  |  Main: {h['profit_main']:+.2f}€  |  Cover: {h['profit_cover']:+.2f}€{arb}"
                break

        data = [
            (rank,                                      "center", True),
            (str(row.get("Date",   "")),                "center", False),
            (str(row.get("Time",   "")),                "center", False),
            (str(row.get("League", "")),                "left",   False),
            (str(row.get("Match",  "")),                "left",   True),
            (str(row.get("Market", "")),                "center", True),
            (round(row["Odds"], 2),                     "center", True),
            (str(row.get("Type",   "")),                "center", False),
            (str(row.get("Stat %", "")),                "center", False),
            (str(row.get("Book %", row.get("Sharp %","—"))), "center", False),
            (round(row["EV"], 4),                       "center", True),
            (row.get("λ", "") if pd.notna(row.get("λ")) else "", "center", False),
            (row.get("μ", "") if pd.notna(row.get("μ")) else "", "center", False),
            (row.get("Home Adv", "") if pd.notna(row.get("Home Adv")) else "", "center", False),
            (row.get("H PPG", "") if pd.notna(row.get("H PPG")) else "", "center", False),
            (row.get("A PPG", "") if pd.notna(row.get("A PPG")) else "", "center", False),
            (str(row.get("Confidence", "")),            "center", False),
            (str(row.get("Agreement",  "")),            "center", False),
            (str(row.get("Trend",      "")),            "center", False),
            (f"{score:.3f}",                            "center", True),
            (g,                                         "center", True),
            (cov_txt,                                   "left",   False),
            (hedge_txt,                                 "left",   False),
        ]
        for col, (val, align, bold) in enumerate(data, 1):
            _c(ws1, r, col, val, bold=bold, bg=row_bg, align=align, sz=9, wrap=(col >= 22))

        sc = ws1.cell(row=r, column=20)
        sc.font  = Font(bold=True, color=_score_fg(score), size=9, name="Calibri")
        sc.fill  = PatternFill("solid", fgColor=_score_bg(score))

        gc = ws1.cell(row=r, column=21)
        grade_map = {"S":(C_GREEN_DK,C_GREEN_BG),"A":(C_NAVY,C_NAVY_LT),
                     "B":(C_AMBER_DK,C_AMBER_BG),"C":(C_RED_DK,C_RED_BG)}
        gfg, gbg = grade_map.get(g, (C_BLACK, C_WHITE))
        gc.font  = Font(bold=True, color=gfg, size=9, name="Calibri")
        gc.fill  = PatternFill("solid", fgColor=gbg)

    last_r = len(selected) + 5
    ws1.row_dimensions[last_r].height = 18
    ws1.merge_cells(f"A{last_r}:D{last_r}")
    _c(ws1, last_r, 1, "AVERAGES", bold=True, bg=C_NAVY_LT, fg=C_NAVY, align="center", sz=9)
    _c(ws1, last_r, 7, f"=AVERAGE(G5:G{last_r-1})", bold=True, bg=C_NAVY_LT, fg=C_NAVY, sz=9)
    _c(ws1, last_r,11, f"=AVERAGE(K5:K{last_r-1})", bold=True, bg=C_NAVY_LT, fg=C_NAVY, sz=9)
    ws1.cell(row=last_r, column=7).number_format  = "0.00"
    ws1.cell(row=last_r, column=11).number_format = "0.000"
    ws1.print_area = f"A1:{get_column_letter(n_cols)}{last_r}"

    # Sheet 2: COVERAGE ANALYSIS
    ws2 = wb.create_sheet("Coverage Analysis")
    _set_tab(ws2, C_ORANGE)
    ws2.sheet_view.showGridLines = False
    ws2.row_dimensions[1].height = 28
    ws2.row_dimensions[2].height = 22

    ws2.merge_cells("A1:J1")
    t2 = ws2["A1"]
    t2.value     = f"  COVERAGE & HEDGE  ·  High Odds Picks (≥ {HIGH_ODDS_THRESHOLD})"
    t2.font      = Font(bold=True, size=12, color=C_WHITE, name="Calibri")
    t2.fill      = PatternFill("solid", fgColor=C_ORANGE)
    t2.alignment = Alignment(horizontal="left", vertical="center")

    hdrs2 = ["Match","Main Market","Main Odds","Cover Market","Cover Odds","Source","Main Stake €10","Cover Stake","If Main Wins","If Cover Wins"]
    cws2  = [32,14,11,16,11,14,13,12,14,14]
    for col,(h,w) in enumerate(zip(hdrs2,cws2),1):
        _h(ws2, 2, col, h, bg=C_ORANGE, sz=9)
        ws2.column_dimensions[get_column_letter(col)].width = w

    r2 = 3
    high_picks = selected[selected["Odds"] >= HIGH_ODDS_THRESHOLD]
    if high_picks.empty:
        ws2.merge_cells("A3:J3")
        _c(ws2, 3, 1, f"No high-odds picks (≥ {HIGH_ODDS_THRESHOLD}) in this selection.", align="center", fg="595959")
    else:
        for _, row in high_picks.iterrows():
            covs  = find_coverage(row, all_picks)
            first = True
            if not covs:
                bg = _alt(r2)
                _c(ws2,r2,1,row.get("Match",""),  align="left", bg=bg, bold=True)
                _c(ws2,r2,2,row.get("Market",""), align="center", bg=bg)
                _c(ws2,r2,3,row["Odds"],           align="center", bg=bg)
                for col in range(4,11): _c(ws2,r2,col,"—",align="center",bg=bg,fg="595959")
                r2 += 1
                continue
            for cv in covs:
                co    = cv["coverage_odds"]
                is_r  = co and isinstance(co, float)
                h_res = calc_hedge(row["Odds"], co, 10.0) if is_r else {}
                is_arb= h_res.get("is_arb", False)
                bg    = C_ARB_BG if is_arb else (C_COVER_BG if first else _alt(r2))
                src   = "✅ Excel" if cv["source"]=="excel" else "💡 Suggested"
                _c(ws2,r2,1,row.get("Match","") if first else "", align="left", bg=bg, bold=first)
                _c(ws2,r2,2,row.get("Market","") if first else "", align="center", bg=bg)
                _c(ws2,r2,3,row["Odds"] if first else "", align="center", bg=bg)
                _c(ws2,r2,4,cv["coverage_market"], align="center", bg=bg, bold=True)
                _c(ws2,r2,5,co if is_r else "—",   align="center", bg=bg)
                _c(ws2,r2,6,src+(" ✅ARB" if is_arb else ""), align="center", bg=bg,
                   fg=C_GREEN_DK if is_arb else (C_NAVY if cv["source"]=="excel" else "595959"))
                _c(ws2,r2,7,"€10", align="center", bg=bg)
                _c(ws2,r2,8,f"€{h_res['cover_stake']}" if h_res else "—", align="center", bg=bg)
                for col, val in [(9,h_res.get("profit_main")), (10,h_res.get("profit_cover"))]:
                    if val is not None:
                        _c(ws2,r2,col,f"{val:+.2f}€",align="center",bg=bg,bold=True,
                           fg=C_GREEN_DK if val>0 else C_RED_DK)
                    else:
                        _c(ws2,r2,col,"—",align="center",bg=bg,fg="595959")
                r2 += 1; first = False
            for col in range(1,11):
                ws2.cell(row=r2,column=col).fill   = PatternFill("solid",fgColor=C_GRAY_MID)
                ws2.cell(row=r2,column=col).border = _B
            ws2.row_dimensions[r2].height = 4
            r2 += 1

    # Sheet 3: SCORE BREAKDOWN (now includes Threshold Detail)
    ws3 = wb.create_sheet("Score Breakdown")
    _set_tab(ws3, C_GREEN_DK)
    ws3.sheet_view.showGridLines = False
    ws3.row_dimensions[1].height = 26
    ws3.row_dimensions[2].height = 22

    ws3.merge_cells("A1:L1")
    t3 = ws3["A1"]
    t3.value     = "  SCORE BREAKDOWN  ·  How each pick was scored (incl. Model Strength & Threshold Bonus)"
    t3.font      = Font(bold=True, size=12, color=C_WHITE, name="Calibri")
    t3.fill      = PatternFill("solid", fgColor=C_GREEN_DK)
    t3.alignment = Alignment(horizontal="left", vertical="center")

    hdrs3 = ["#","Match","Market","Odds","Type",
             f"EV ({int(WEIGHT_EV*100)}%)",
             f"Conf ({int(WEIGHT_CONFIDENCE*100)}%)",
             f"Stat ({int(WEIGHT_STAT_P*100)}%)",
             f"Odds ({int(WEIGHT_ODDS*100)}%)",
             f"Trend ({int(WEIGHT_TREND*100)}%)",
             f"Model ({int(WEIGHT_MODEL*100)}%)",
             "Thresh", "Threshold Detail", "Penalty", "TOTAL"]
    cws3 = [4,30,10,8,14,10,10,10,10,10,10,8,30,9,10]
    for col,(h,w) in enumerate(zip(hdrs3,cws3),1):
        _h(ws3, 2, col, h, bg=C_GREEN_DK, sz=9)
        ws3.column_dimensions[get_column_letter(col)].width = w

    for rank,(_, row) in enumerate(selected.iterrows(), 1):
        r3    = rank + 2
        ws3.row_dimensions[r3].height = 17
        score = float(row.get("Composite_Score", 0))
        bg    = _alt(rank)
        ev_s  = min(1.0,max(0.0,(row["EV"]-1.0)/0.15))
        emp_c = row.get("Agreement_num", row["Confidence_num"])
        conf_s= emp_c * row.get("Agreement_num", 0.5)
        stat_s= min(1.0,max(0.0,(row["StatP_num"]-0.40)/0.40))
        odds_s= math.exp(-0.5*((row["Odds"]-ODDS_SWEET_SPOT)/ODDS_SIGMA)**2)
        trend_s= row["Trend_score"]
        model_s = compute_model_strength(row)
        penalty= float(row.get("Penalty", 0.0))
        thresh_bonus = row.get("Threshold_Bonus", 0.0)
        thresh_detail = row.get("Threshold_Detail", "None")
        # Αν δεν υπάρχει detail, δείχνουμε τα κριτήρια της αγοράς
        if thresh_bonus == 0.0 and thresh_detail == "None":
            market = row.get("Market", "")
            thresh_detail = format_threshold_criteria(market, optimal_thresholds)
        
        vals = [rank, str(row.get("Match","")), str(row.get("Market","")), round(row["Odds"],2), str(row.get("Type","")),
                f"{ev_s*WEIGHT_EV:.3f}", f"{conf_s*WEIGHT_CONFIDENCE:.3f}",
                f"{stat_s*WEIGHT_STAT_P:.3f}", f"{odds_s*WEIGHT_ODDS:.3f}",
                f"{trend_s*WEIGHT_TREND:.3f}", f"{model_s*WEIGHT_MODEL:.3f}",
                f"{thresh_bonus:.3f}", thresh_detail,
                f"{-penalty:.3f}", f"{score:.3f}"]
        for col,(v) in enumerate(vals,1):
            _c(ws3, r3, col, v, bold=(col==14), bg=bg, sz=9)
        tc = ws3.cell(row=r3, column=14)
        tc.font = Font(bold=True, color=_score_fg(score), size=9, name="Calibri")
        tc.fill = PatternFill("solid", fgColor=_score_bg(score))

    # Sheet 4: THRESHOLD INFO
    ws5 = wb.create_sheet("Threshold Info")
    _set_tab(ws5, C_ORANGE)
    ws5.sheet_view.showGridLines = False
    ws5.column_dimensions["A"].width = 20
    ws5.column_dimensions["B"].width = 60
    ws5.row_dimensions[1].height = 28
    ws5.merge_cells("A1:B1")
    t5 = ws5["A1"]
    t5.value = "  THRESHOLD CRITERIA (Learned from Ledger)"
    t5.font = Font(bold=True, size=13, color=C_WHITE, name="Calibri")
    t5.fill = PatternFill("solid", fgColor=C_ORANGE)
    t5.alignment = Alignment(horizontal="center", vertical="center")

    r5 = 2
    if optimal_thresholds:
        for market, criteria in optimal_thresholds.items():
            ws5.merge_cells(f"A{r5}:B{r5}")
            sc = ws5[f"A{r5}"]
            sc.value = f"  Market: {market}"
            sc.font = Font(bold=True, size=11, color=C_WHITE, name="Calibri")
            sc.fill = PatternFill("solid", fgColor=C_NAVY_MID)
            ws5.row_dimensions[r5].height = 22
            r5 += 1
            for feat, (thresh, direction) in criteria.items():
                short_name = feat.replace(" (Lambda)", "λ").replace(" (Mu)", "μ").replace("Home_Adv", "HA").replace("H_PPG", "HPPG").replace("A_PPG", "APPG")
                symbol = "≥" if direction == "above" else "<"
                desc = f"{short_name} {symbol} {thresh:.3f}"
                _c(ws5, r5, 1, desc, bold=True, align="left", bg=C_NAVY_LT)
                _c(ws5, r5, 2, f"Favours picks where {short_name} is {'higher' if direction=='above' else 'lower'} than {thresh:.3f}", align="left")
                ws5.row_dimensions[r5].height = 18
                r5 += 1
            r5 += 1
    else:
        _c(ws5, 2, 1, "No thresholds found yet. Need more data per market.", align="left")

    # Sheet 5: GUIDE
    ws4 = wb.create_sheet("Guide")
    _set_tab(ws4, "BF8F00")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 26
    ws4.column_dimensions["B"].width = 70
    ws4.row_dimensions[1].height = 28

    ws4.merge_cells("A1:B1")
    tg = ws4["A1"]
    tg.value     = "  ORACLE ANALYST V6  ·  User Guide"
    tg.font      = Font(bold=True, size=13, color=C_WHITE, name="Calibri")
    tg.fill      = PatternFill("solid", fgColor="BF8F00")
    tg.alignment = Alignment(horizontal="center", vertical="center")

    guide = [
        ("V6 IMPROVEMENTS", [
            ("Model Strength Score", "Evaluates λ, μ, Home Adv, H PPG, A PPG for the specific market."),
            ("Empirical confidence",  "Win rates from historical ledger replace subjective confidence scores."),
            ("Ledger-adjusted weights","Type bonuses auto-calibrate from actual VALUE/PATTERN/TRAP performance."),
            ("Correlation penalty",   "Picks in same match with high historical correlation get score reduction."),
            ("Threshold Bonus",       "Auto-learns winning conditions (e.g. low μ for Under 2.5) and boosts picks that meet them."),
        ]),
        ("GRADES", [
            ("S  (≥ 0.75)", "Diamond. All factors align."),
            ("A  (0.60–0.74)", "Excellent."),
            ("B  (0.45–0.59)", "Solid. Reduce stake 25%."),
            ("C  (< 0.45)", "Borderline. Consider skipping."),
        ]),
        ("COVERAGE", [
            ("✅ Excel", "Found in Oracle output — odds ready."),
            ("💡 Suggested", "Theoretical — search your bookmaker."),
            ("✅ ARB", "Guaranteed profit on both outcomes."),
        ]),
    ]
    r4 = 2
    for section, items in guide:
        ws4.merge_cells(f"A{r4}:B{r4}")
        sc = ws4[f"A{r4}"]
        sc.value     = f"  {section}"
        sc.font      = Font(bold=True, size=10, color=C_WHITE, name="Calibri")
        sc.fill      = PatternFill("solid", fgColor=C_NAVY_MID)
        sc.alignment = Alignment(horizontal="left", vertical="center")
        ws4.row_dimensions[r4].height = 20
        r4 += 1
        for term, desc in items:
            ws4.row_dimensions[r4].height = 18
            ka = ws4.cell(row=r4, column=1, value=term)
            ka.font      = Font(bold=True, size=9, color=C_NAVY, name="Calibri")
            ka.fill      = PatternFill("solid", fgColor=C_NAVY_LT)
            ka.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ka.border    = _B
            vb = ws4.cell(row=r4, column=2, value=desc)
            vb.font      = Font(size=9, color="595959", name="Calibri")
            vb.fill      = PatternFill("solid", fgColor=C_WHITE)
            vb.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            vb.border    = _B
            r4 += 1
        r4 += 1

    wb.save(path)
    print(f"✅  Report → {path}  ({len(selected)} picks, 5 sheets)")

# =====================================================================
# MAIN
# =====================================================================
def main():
    print("="*70)
    print("  ORACLE ANALYST V6 — DYNAMIC THRESHOLD BONUS")
    print("  (Auto-learns from Ledger, boosts winning conditions)")
    print("="*70)
    
    ledger = load_ledger()
    if not ledger.empty:
        print(f"📊 Loaded {len(ledger)} settled picks from ledger")
        market_rates, type_rates, league_rates, odds_rates = compute_empirical_rates(ledger)
        global WEIGHT_EV, WEIGHT_CONFIDENCE, WEIGHT_STAT_P, WEIGHT_ODDS, WEIGHT_TREND, WEIGHT_MODEL, TYPE_BONUS
        (WEIGHT_EV, WEIGHT_CONFIDENCE, WEIGHT_STAT_P, WEIGHT_ODDS, 
         WEIGHT_TREND, WEIGHT_MODEL, TYPE_BONUS) = adjust_weights_from_ledger(ledger)
        corr_dict = compute_market_correlation(ledger)
        print(f"   Market correlations computed: {len(corr_dict)} pairs")
        
        print("🧠 Running Threshold Analysis on Ledger...")
        optimal_thresholds = find_optimal_thresholds(ledger)
        print(f"   Found thresholds for {len(optimal_thresholds)} markets.")
        for market, criteria in optimal_thresholds.items():
            print(f"      {market}: {format_threshold_criteria(market, optimal_thresholds)}")
    else:
        print("⚠️ No historical ledger found. Using default weights and no threshold bonus.")
        market_rates, type_rates, league_rates, odds_rates = {}, {}, {}, {}
        corr_dict = {}
        optimal_thresholds = {}
    
    excel_path = find_oracle_excel()
    all_picks = load_picks(excel_path)
    if all_picks.empty:
        print("❌ No picks found.")
        return
    
    weights = (WEIGHT_EV, WEIGHT_CONFIDENCE, WEIGHT_STAT_P, WEIGHT_ODDS, WEIGHT_TREND, WEIGHT_MODEL)
    selected = select_top_picks(all_picks, market_rates, type_rates, league_rates, odds_rates, weights, corr_dict, optimal_thresholds)
    
    print(f"\n📊 TOP {len(selected)} PICKS (incl. Threshold Bonus):")
    for rank, (_, row) in enumerate(selected.iterrows(), 1):
        score = row['Composite_Score']
        grade = "S" if score>=0.75 else "A" if score>=0.60 else "B" if score>=0.45 else "C"
        print(f"{rank:2}. {str(row['Match'])[:32]:32} {row['Market']:10} {row['Odds']:5.2f} EV:{row['EV']:.3f} Score:{score:.3f} {grade}")
    
    write_report(selected, all_picks, OUTPUT_FILE, optimal_thresholds)
    print(f"\n📂 Report saved: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
